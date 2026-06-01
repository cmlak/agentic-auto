import os
import json
import openpyxl
from decimal import Decimal
from openpyxl.styles import Font, PatternFill
from django.core.management.base import BaseCommand
from pydantic import BaseModel, Field
from typing import List, Literal
from google import genai
from google.genai import types
from tenacity import retry, stop_after_attempt, wait_exponential

# ==========================================
# 1. PYDANTIC SCHEMAS FOR AI
# ==========================================
class MatchedPair(BaseModel):
    agentic_ids: List[str] = Field(..., description="List of Agentic IDs (e.g., ['A-004', 'A-005'])")
    balancika_ids: List[str] = Field(..., description="List of Balancika IDs (e.g., ['B-005'])")
    variance_usd: float = Field(..., description="Calculate: (Sum of Agentic amounts) - (Sum of Balancika amounts). Must be exactly 0.0 if perfect match.")
    # 💡 ENHANCEMENT: Forced detailed logging in the schema description
    match_reason: str = Field(..., description="Explain the logic tying these together. REQUIRED: You MUST explicitly quote the Date, Amount, and Description from BOTH systems in this explanation. Also explain the exact reason for any variance_usd.")

class OrphanRecord(BaseModel):
    system: Literal["Agentic", "Balancika"]
    record_id: str = Field(..., description="The ID of the unmatched record.")
    # 💡 ENHANCEMENT: Forced detailed logging for orphans
    omission_reason: str = Field(..., description="Explain why this record has no counterpart. REQUIRED: You MUST explicitly quote the Date, Amount, and Description of this orphaned record in your explanation.")

class ReconciliationResult(BaseModel):
    reasoning: str = Field(..., description="Explain your methodology for matching these complex leftovers.")
    semantic_matches: List[MatchedPair] = []
    discrepancies: List[OrphanRecord] = []

# ==========================================
# 2. DJANGO MANAGEMENT COMMAND
# ==========================================
class Command(BaseCommand):
    help = 'Hybrid Python + AI Cross-check of Trade Payables'

    def handle(self, *args, **options):
        api_key = os.environ.get("GEMINI_API_KEY_2") 
        if not api_key:
            self.stdout.write(self.style.ERROR("GEMINI_API_KEY_2 environment variable not found."))
            return

        self.client = genai.Client(api_key=api_key)
        self.AUDIT_MODEL = "gemini-3.1-pro-preview" # or gemini-2.5-pro
        
        base_url = r'C:\bakertilly\BakerTilly\CCKT\Balancika\Jan cross check'
        agentic_filename = 'agentic_CIP_march.xlsx'
        balancika_filename = 'balancika_CIP_march.xlsx'
        report_filename = 'CIP_Discrepancy_Report.xlsx'
        
        agentic_file = os.path.join(base_url, agentic_filename)
        balancika_file = os.path.join(base_url, balancika_filename)
        report_file = os.path.join(base_url, report_filename)

        self.stdout.write("Loading datasets...")
        agentic_records = self.load_xlsx(agentic_file, system='Agentic', prefix='A')
        balancika_records = self.load_xlsx(balancika_file, system='Balancika', prefix='B')
        
        # 1. DETERMINISTIC PYTHON MATCHING (Math First)
        self.stdout.write("Running pure programmatic math reconciliation...")
        exact_matches = self.programmatic_exact_match(agentic_records, balancika_records)
        
        # Isolate the leftovers for the AI
        agentic_orphans = {k: v for k, v in agentic_records.items() if not v['matched']}
        balancika_orphans = {k: v for k, v in balancika_records.items() if not v['matched']}
        
        self.stdout.write(self.style.SUCCESS(f"Programmatic pass cleared {len(exact_matches)} pairs."))
        self.stdout.write(f"Leftovers for AI Analysis: {len(agentic_orphans)} Agentic, {len(balancika_orphans)} Balancika.")

        # 2. SEMANTIC AI MATCHING (Text & Split Invoices)
        ai_result = None
        if agentic_orphans or balancika_orphans:
            self.stdout.write(f"Initiating AI semantic reconciliation via {self.AUDIT_MODEL}...\n")
            try:
                ai_result = self.run_ai_reconciliation(agentic_orphans, balancika_orphans)
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"AI Reconciliation failed: {str(e)}"))
                return
        
        # 3. CONSOLIDATE AND EXPORT
        self.process_and_report(exact_matches, ai_result, agentic_records, balancika_records, report_file)

    def load_xlsx(self, filepath, system, prefix):
        records = {}
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip() if cell.value else f"Col_{i}" for i, cell in enumerate(sheet[1])]
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
            row_dict = dict(zip(headers, row))
            date_val = row_dict.get('Date')
            if not date_val: continue
            
            date_str = str(date_val).split(' ')[0].strip()
            
            try:
                debit = float(str(row_dict.get('Debit', 0)).replace(',', '') or 0)
                credit = float(str(row_dict.get('Credit', 0)).replace(',', '') or 0)
            except (ValueError, TypeError):
                continue
                
            if debit == 0 and credit == 0: continue
            
            desc = str(row_dict.get('Description') or '').strip()
            
            if system == 'Balancika' and not desc:
                desc = str(row_dict.get('Vendor / Customer / Employee') or row_dict.get('Source') or '').strip()
                
            if system == 'Agentic':
                source = str(row_dict.get('Source') or '').strip()
                if source and source.lower() != 'none':
                    desc = f"{desc} | Source: {source}"

            record_id = f"{prefix}-{idx:03d}"
            records[record_id] = {
                'id': record_id,
                'date': date_str,
                'desc': desc,
                'debit': debit,
                'credit': credit,
                'matched': False
            }
        return records

    def programmatic_exact_match(self, agentic, balancika):
        """Pure Python algorithm to clear exact mathematical matches first."""
        matches = []
        for a_id, a_rec in agentic.items():
            if a_rec['matched']: continue
            
            for b_id, b_rec in balancika.items():
                if b_rec['matched']: continue
                
                if a_rec['debit'] == b_rec['debit'] and a_rec['credit'] == b_rec['credit']:
                    a_rec['matched'] = True
                    b_rec['matched'] = True
                    matches.append({
                        'agentic_ids': [a_id],
                        'balancika_ids': [b_id],
                        'match_reason': f"Programmatic exact amount match ({a_rec['credit']} CR, {a_rec['debit']} DR)"
                    })
                    break
        return matches

    @retry(stop=stop_after_attempt(4), wait=wait_exponential(multiplier=2, min=3, max=15), reraise=True)
    def run_ai_reconciliation(self, agentic_orphans, balancika_orphans):
        a_payload = json.dumps([{k: v for k, v in r.items() if k != 'matched'} for r in agentic_orphans.values()], indent=2)
        b_payload = json.dumps([{k: v for k, v in r.items() if k != 'matched'} for r in balancika_orphans.values()], indent=2)

        prompt = f"""
        TASK: You are an elite Forensic Accountant. An algorithmic pass has cleared all exact mathematical matches. 
        You must reconcile the complex LEFTOVERS. 
        
        AGENTIC ORPHANS:
        {a_payload}
        
        BALANCIKA ORPHANS:
        {b_payload}
        
        FORENSIC INSTRUCTIONS:
        1. SPLIT INVOICES: One Balancika record might equal the sum of multiple Agentic records (or vice versa).
        2. HIDDEN VARIANCES: Look for records with matching text/dates but slight amount differences. A discrepancy (e.g., $7.50, $15.00) is usually caused by embedded Bank Transfer Fees, WHT (Withholding Tax), VAT, or Currency Exchange roundings.
        3. STRICT MATH: For EVERY semantic match you make, you MUST calculate the `variance_usd`. If there is a variance, hypothesize what it represents.
        4. DETAILED LOGGING (CRITICAL): Your `match_reason` and `omission_reason` MUST explicitly cite the Date, Debit/Credit Amount, and Description of the records involved. Do not write vague summaries. (Example of Good Output: "Agentic [2026-03-24: $50.00 for 'Steel'] matches Balancika [2026-03-24: $57.50 for 'Construction Steel'] with a $7.50 variance likely due to delivery fees.")
        5. ORPHANS: If a variance is too large to logically explain, do not force a match. Output them as Discrepancies and cite their full date/amount/description details.
        6. EXHAUSTIVE: Account for EVERY SINGLE ID listed above.
        """

        response = self.client.models.generate_content(
            model=self.AUDIT_MODEL,
            contents=[prompt],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                response_schema=ReconciliationResult,
                temperature=0.0 
            )
        )
        
        result = response.parsed
        
        # --- Bi-Directional Anti-Laziness Check ---
        ai_handled_a = sum(len(m.agentic_ids) for m in result.semantic_matches) + len([o for o in result.discrepancies if o.system == "Agentic"])
        ai_handled_b = sum(len(m.balancika_ids) for m in result.semantic_matches) + len([o for o in result.discrepancies if o.system == "Balancika"])
        
        if ai_handled_a != len(agentic_orphans) or ai_handled_b != len(balancika_orphans):
            raise ValueError(f"AI Laziness Detected: Missing IDs in output. Expected A:{len(agentic_orphans)}, B:{len(balancika_orphans)}. Handled A:{ai_handled_a}, B:{ai_handled_b}.")
            
        return result

    def process_and_report(self, exact_matches, ai_result, agentic_db, balancika_db, report_file):
        agentic_diff = Decimal('0.00')
        balancika_diff = Decimal('0.00')
        discrepancy_rows = []

        all_matches = exact_matches.copy()
        
        if ai_result:
            for sm in ai_result.semantic_matches:
                a_sum = sum(Decimal(str(agentic_db[a]['credit'])) - Decimal(str(agentic_db[a]['debit'])) for a in sm.agentic_ids)
                b_sum = sum(Decimal(str(balancika_db[b]['credit'])) - Decimal(str(balancika_db[b]['debit'])) for b in sm.balancika_ids)
                
                variance = a_sum - b_sum
                
                if variance != Decimal('0.00'):
                    agentic_diff += a_sum
                    balancika_diff += b_sum
                    
                    ids_involved = f"Agentic: {', '.join(sm.agentic_ids)} | Balancika: {', '.join(sm.balancika_ids)}"
                    discrepancy_rows.append([
                        'FUZZY MATCH VARIANCE', ids_involved, 'N/A', sm.match_reason, 
                        'N/A', 'N/A', f"Internal Variance of ${variance}"
                    ])

                all_matches.append({
                    'agentic_ids': sm.agentic_ids,
                    'balancika_ids': sm.balancika_ids,
                    'match_reason': f"[AI SEMANTIC] {sm.match_reason} (Variance: ${variance})"
                })

            for orphan in ai_result.discrepancies:
                if orphan.system == 'Agentic':
                    rec = agentic_db.get(orphan.record_id)
                    if rec:
                        agentic_diff += Decimal(str(rec['credit'])) - Decimal(str(rec['debit']))
                        discrepancy_rows.append(['Agentic', rec['id'], rec['date'], rec['desc'], rec['debit'], rec['credit'], orphan.omission_reason])
                else:
                    rec = balancika_db.get(orphan.record_id)
                    if rec:
                        balancika_diff += Decimal(str(rec['credit'])) - Decimal(str(rec['debit']))
                        discrepancy_rows.append(['Balancika', rec['id'], rec['date'], rec['desc'], rec['debit'], rec['credit'], orphan.omission_reason])

        net_difference = agentic_diff - balancika_diff
        
        self.stdout.write(self.style.SUCCESS(f"\nTotal Exact Matches: {len(exact_matches)}"))
        if ai_result:
            self.stdout.write(self.style.SUCCESS(f"Total AI Semantic/Split Matches: {len(ai_result.semantic_matches)}"))
            
        self.stdout.write(f"Net Period Transaction Variance (Agentic): ${agentic_diff}")
        self.stdout.write(f"Net Period Transaction Variance (Balancika): ${balancika_diff}")
        self.stdout.write(self.style.ERROR(f"Total Period Transaction Discrepancy: ${net_difference}"))
        
        self.stdout.write(self.style.WARNING("\nNote: This variance only reflects transactions during the parsed period. If there is an overall ledger difference, check the Opening Balances for the period."))

        # --- EXCEL EXPORT ---
        wb = openpyxl.Workbook()
        ws_disc = wb.active
        ws_disc.title = "Discrepancies"
        
        ws_disc.append(["Reconciliation Summary"])
        ws_disc["A1"].font = Font(bold=True, size=14)
        ws_disc.append(["Net transaction variance originating from Agentic", float(agentic_diff)])
        ws_disc.append(["Net transaction variance originating from Balancika", float(balancika_diff)])
        ws_disc.append(["Total Isolated Period Discrepancy", float(net_difference)])
        ws_disc["A4"].font = Font(bold=True)
        ws_disc["B4"].font = Font(bold=True)
        ws_disc.append([]) 
        
        headers = ['Source System', 'Record ID', 'Date', 'Description', 'Debit', 'Credit', 'AI Reason for Omission/Variance']
        ws_disc.append(headers)
        
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws_disc[6]:
            cell.fill = header_fill
            cell.font = header_font
            
        for row in discrepancy_rows:
            ws_disc.append(row)
            
        for col, width in {'A': 20, 'B': 25, 'C': 12, 'D': 50, 'E': 10, 'F': 10, 'G': 100}.items():
            ws_disc.column_dimensions[col].width = width

        ws_match = wb.create_sheet(title="All Matched Pairs")
        ws_match.append(['Agentic IDs', 'Balancika IDs', 'Match Reasoning'])
        
        for cell in ws_match[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        for match in all_matches:
            a_str = ", ".join(match['agentic_ids'])
            b_str = ", ".join(match['balancika_ids'])
            ws_match.append([a_str, b_str, match['match_reason']])
            
        ws_match.column_dimensions['A'].width = 20
        ws_match.column_dimensions['B'].width = 20
        ws_match.column_dimensions['C'].width = 100

        wb.save(report_file)
        self.stdout.write(self.style.SUCCESS(f"Reconciliation complete. Report saved to {report_file}"))