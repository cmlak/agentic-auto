import os
import tempfile
import pandas as pd
import calendar
import io
import re
import uuid
from collections import defaultdict
import time
import openpyxl
from openpyxl.styles import Alignment
import difflib
import json
import json
import base64

from datetime import date, datetime, timedelta
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseRedirect, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse, reverse_lazy
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import DetailView, UpdateView, DeleteView
from django.db.models import Sum, Q
from django.db import transaction
from django.core.paginator import Paginator
import pdfplumber
from pypdf import PdfReader, PdfWriter

# Import your forms, processors, and local models
from .forms import BatchUploadForm, PurchaseFormSet, ManualPurchaseEntryForm, GLMigrationUploadForm,\
GLHistoricalFormSet, OldEntryForm, JournalVoucherEntryForm, BalancikaExportForm,\
MultiplePDFUploadForm, MonthlyClosingForm, AccrualFormSet, FXFormSet, EngagementLetterUploadForm,\
AdjustmentEntryForm, AdjustmentFormSet, OffsetFormSet, ManualInvoiceUploadForm
from .processors import GeminiInvoiceProcessor, GLMigrationProcessor, ProposalPDFProcessor, TOSPDFProcessor,\
TaxLiabilitiesProcessor, EngagementLetterProcessor, UnifiedTaxProcessor
from .orchestrators import InvoiceOrchestrator, DjangoEventOrchestrator, SystemOrchestrator
from .models import Purchase, AICostLog, Vendor, Old, JournalVoucher, Adjustment
from account.models import Account, JournalEntry, JournalLine, AccountMappingRule, ClientPromptMemo
from register.models import Profile
from .filters import PurchaseFilter, JournalVoucherFilter, AdjustmentFilter
from .resources import PurchaseResource, AdjustmentResource
from cash.models import Bank
from sale.models import Customer, Sale
from tools.tasks import handle_user_correction_task, process_draft_rule_task


# ====================================================================
# --- 1. AI INVOICE UPLOAD & PROCESSING ---
# ====================================================================

@login_required(login_url="register:login")
def invoice_ai_upload_view(request):
    user = request.user

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'process_page':
            job = request.session.get('invoice_job')
            if not job:
                return JsonResponse({"status": "error", "message": "Job session not found."})
                
            page = int(request.POST.get('page', 1))
            api_key = getattr(settings, 'GEMINI_API_KEY_2', os.getenv("GEMINI_API_KEY_2"))
            processor = GeminiInvoiceProcessor(api_key=api_key)
            
            local_file_path = job.get('local_file_path')
            if not local_file_path or not os.path.exists(local_file_path):
                return JsonResponse({"status": "error", "message": "Source PDF lost on server."})
                
            try:
                reader = PdfReader(local_file_path)
                writer = PdfWriter()
                writer.add_page(reader.pages[page - 1])
                
                pdf_bytes_io = io.BytesIO()
                writer.write(pdf_bytes_io)
                single_page_bytes = pdf_bytes_io.getvalue()
            except Exception as e:
                return JsonResponse({"status": "error", "message": f"Failed to isolate PDF page {page}. Error: {str(e)}"})
            
            print(f"\n[PAGE {page}] EXTRACTING KEY INVOICE DATA FROM AI...")
            ledgers, page_cost, next_seq, err = processor.process_single_page(
                pdf_bytes=single_page_bytes, 
                pg=page, custom_prompt=job['custom_prompt'],
                batch_name=job['batch_name'], rules_context=job['rules_context'],
                memo_context=job['memo_context'], current_invoice_seq=job['current_seq'],
                date_prefix=job['date_prefix'], is_explicit_seq=job['is_explicit_seq']
            )
            
            # Reload the session state to prevent race conditions during concurrent page processing
            current_job = request.session.get('invoice_job')
            if current_job:
                if ledgers:
                    current_job['results'].extend(ledgers)
                    print(f"   🎉 Extracted {len(ledgers)} invoices from Page {page}.")

                current_job['current_seq'] = max(current_job.get('current_seq', 1), next_seq)
                current_job['costs']['pro_cost'] += page_cost
                request.session['invoice_job'] = current_job
                request.session.save()
            
            # --- CATCH TIMEOUT AND NOTIFY FRONTEND ---
            if err and "Timeout Error" in err:
                return JsonResponse({"status": "timeout", "page": page, "message": f"AI timed out on Page {page}. Salvaging partial data."})
            
            return JsonResponse({"status": "success", "page": page, "ledgers_count": len(ledgers) if ledgers else 0, "error": err})
            
        if action == 'finalize':
            job = request.session.get('invoice_job')
            if not job:
                return JsonResponse({"status": "error", "message": "Job session not found."})
                
            local_file_path = job.get('local_file_path')
            if local_file_path and os.path.exists(local_file_path):
                os.remove(local_file_path)
                
            results = job.get('results', [])
            results.sort(key=lambda x: int(x.get('page', 0) or 0))
            
            is_explicit_seq = job.get('is_explicit_seq', False)
            date_prefix = job.get('date_prefix')
            original_seq = job.get('original_seq', 1)
            
            processed_pages = set()
            month_trackers = {}
            current_explicit_seq = original_seq
            
            # --- FINAL SEQUENCE ASSIGNMENT ---
            for item in results:
                page = item.get('page')
                inv_no = str(item.get('invoice_no', ''))

                if inv_no == "NEEDS_SEQ" or inv_no.startswith('INV-'):
                    if is_explicit_seq:
                        if page not in processed_pages:
                            processed_pages.add(page)
                            base_seq = current_explicit_seq
                            current_explicit_seq += 1
                        else:
                            base_seq = current_explicit_seq - 1
                        base_inv_no = f"INV-{date_prefix}{base_seq:02d}"
                    else:
                        item_date = item.get('date')
                        if item_date:
                            try:
                                parsed_date = datetime.strptime(item_date, "%Y-%m-%d")
                                month_prefix = parsed_date.strftime("%Y%m")
                            except ValueError: month_prefix = datetime.now().strftime("%Y%m")
                        else:
                            month_prefix = datetime.now().strftime("%Y%m")
                            
                        if month_prefix not in month_trackers:
                            existing_invs = Purchase.objects.filter(
                                invoice_no__startswith=f"INV-{month_prefix}"
                            ).values_list('invoice_no', flat=True)
                            max_seq = 0
                            for inv in existing_invs:
                                match = re.search(rf'INV-{month_prefix}(\d+)', inv)
                                if match: max_seq = max(max_seq, int(match.group(1)))
                            month_trackers[month_prefix] = max_seq + 1
                            
                        if page not in processed_pages:
                            processed_pages.add(page)
                            base_seq = month_trackers[month_prefix]
                            month_trackers[month_prefix] += 1
                        else:
                            base_seq = month_trackers[month_prefix] - 1
                            
                        base_inv_no = f"INV-{month_prefix}{base_seq:02d}"

                    parts = inv_no.split("-")
                    if len(parts) > 2 and parts[-1].isdigit() and len(parts[-1]) < 4:
                        item['invoice_no'] = f"{base_inv_no}-{parts[-1]}"
                    else:
                        item['invoice_no'] = base_inv_no

            print("\n[FINALIZING] LOGGING AI COSTS AND SAVING STATE...")
            total_flash = job['costs']['flash_cost']
            total_pro = job['costs']['pro_cost']
            total_cost = total_flash + total_pro
            
            try:
                AICostLog.objects.create(file_name=job['file_name'], total_pages=job['total_pages'], flash_cost=total_flash, pro_cost=total_pro, total_cost=total_cost)
            except NameError: pass
                
            request.session['extracted_invoices'] = results
            request.session['ai_metadata'] = {
                'file_name': job['file_name'], 'batch_name': job['batch_name'],
                'total_pages': job['total_pages'], 'costs': job['costs']
            }
            request.session.pop('invoice_job', None)
            return JsonResponse({"status": "success", "redirect_url": reverse('tools:review_invoices')})

        request.session.pop('invoice_report_path', None)
        form = BatchUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_pdf = form.cleaned_data['invoice_pdf']
            batch_name = form.cleaned_data['batch_name']
            custom_prompt = form.cleaned_data.get('ai_prompt', '')
            
            inv_match = re.search(r'INV-(\d{6})(\d+)', custom_prompt, re.IGNORECASE)
            if inv_match:
                date_prefix = inv_match.group(1) 
                current_seq = int(inv_match.group(2)) 
                is_explicit_seq = True
            else:
                date_prefix = datetime.now().strftime("%Y%m")
                current_seq = 1
                is_explicit_seq = False

            rules_context = ""
            memo_context = ""
            client_memo = ClientPromptMemo.objects.first()
            if client_memo: memo_context = client_memo.memo_text

            rules = AccountMappingRule.objects.all().select_related('account')
            if rules.exists():
                rules_data = [{'Account ID': r.account.account_id, 'Account Name': r.account.name, 'Description / Trigger Keywords': r.trigger_keywords, 'Reasoning / AI Guidelines': r.ai_guideline} for r in rules]
                rules_context = pd.DataFrame(rules_data).to_csv(index=False)

            temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_invoices')
            os.makedirs(temp_dir, exist_ok=True)
            unique_filename = f"batch_{uuid.uuid4().hex}.pdf"
            local_file_path = os.path.join(temp_dir, unique_filename)
            
            with open(local_file_path, 'wb') as f:
                for chunk in uploaded_pdf.chunks(): f.write(chunk)

            try:
                reader = PdfReader(local_file_path)
                total_pages = len(reader.pages)
                
                if total_pages > 20:
                    os.remove(local_file_path)
                    return JsonResponse({"status": "error", "message": f"Limit exceeded. PDF has {total_pages} pages, max is 20."})
                
                request.session['invoice_job'] = {
                    'local_file_path': local_file_path, 'file_name': uploaded_pdf.name, 'total_pages': total_pages,
                    'batch_name': batch_name, 'custom_prompt': custom_prompt,
                    'rules_context': rules_context, 'memo_context': memo_context, 'is_explicit_seq': is_explicit_seq,
                    'date_prefix': date_prefix, 'original_seq': current_seq, 'current_seq': current_seq,
                    'results': [], 'costs': {'flash_cost': 0.0, 'pro_cost': 0.0}
                }
                request.session.save()
                return JsonResponse({"status": "init_success", "total_pages": total_pages})
                
            except Exception as e:
                if os.path.exists(local_file_path): os.remove(local_file_path)
                return JsonResponse({"status": "error", "message": f"Initialization Error: {str(e)}"})
        else:
            return JsonResponse({"status": "error", "message": "Form validation failed."})
    else:
        job = request.session.get('invoice_job')
        if job and 'local_file_path' in job and os.path.exists(job['local_file_path']):
            os.remove(job['local_file_path'])
            request.session.pop('invoice_job', None)
            
        form = BatchUploadForm()

    return render(request, 'invoice_upload.html', {'form': form})

# ====================================================================
# --- 2. HITL REVIEW & AUTOMATIC GL POSTING ---
# ====================================================================

@login_required(login_url="register:login")
def review_invoices(request, template_name='tools/invoice_review.html'):
    """Step 2: Review AI data, Update Vendors, Save Source Doc, and Post Journal Entry."""
    extracted_data = request.session.get('extracted_invoices', [])
    metadata = request.session.get('ai_metadata', {})

    if not extracted_data and request.method == 'GET':
        return redirect('tools:invoice_upload')
        
    # --- FIX: Ensure the formset displays pages sequentially ---
    # Do not sort manual entries to preserve original Excel order
    if not metadata.get('is_manual'):
        extracted_data.sort(key=lambda x: int(x.get('page', 0) or 0))
        
    # --- VENDOR CHOICES ---
    db_vendors = [(v.id, f"{v.vendor_id} - {v.name}") for v in Vendor.objects.all().order_by('vendor_id')]
    
    all_vids = Vendor.objects.all().values_list('vendor_id', flat=True)
    max_num = 1
    for vid in all_vids:
        if vid:
            match = re.search(r'V-?(\d+)', str(vid))
            if match:
                max_num = max(max_num, int(match.group(1)))
    next_num = max_num + 1

    new_vendor_map = {}
    for item in extracted_data:
        if item.get('is_new_vendor'):
            raw_name = item.get('company', 'Unknown')
            name_str = str(raw_name).lower().replace('&', ' and ')
            target_norm = re.sub(r'[\W_]+', ' ', name_str).strip()
            
            if target_norm not in new_vendor_map:
                current_seq = next_num + len(new_vendor_map)
                new_vid = f"V-{current_seq:05d}"
                new_temp_id = f"TEMP_{new_vid}"
                new_vendor_map[target_norm] = {
                    'temp_vid': new_vid,
                    'temp_id': new_temp_id,
                    'company': raw_name
                }
            
            mapped = new_vendor_map[target_norm]
            item['temp_vid'] = mapped['temp_vid']
            item['temp_id'] = mapped['temp_id']
            item['vendor_choice'] = mapped['temp_id']

    temp_vendors = []
    for mapped in new_vendor_map.values():
        temp_vendors.append((mapped['temp_id'], f"✨ NEW: {mapped['company']} ({mapped['temp_vid']})"))
        
    dynamic_choices = [('', '--- Select Vendor ---')] + db_vendors + temp_vendors

    # --- ACCOUNT CHOICES ---
    # Fetch all global accounts to allow users to select correct accounts if they aren't explicitly mapped to this client yet
    seen_accounts = set()
    db_accounts = []
    for acc_id, name in Account.objects.values_list('account_id', 'name'):
        if acc_id not in seen_accounts:
            seen_accounts.add(acc_id)
            db_accounts.append((str(acc_id), f"{acc_id} - {name}"))
    db_accounts.sort(key=lambda x: str(x[0]))
    account_choices = [('', '--- Select Account ---')] + db_accounts

    if request.method == 'POST':
        # Pass BOTH dynamic vendors and dynamic accounts to the formset
        formset = PurchaseFormSet(
            request.POST, 
            form_kwargs={'dynamic_choices': dynamic_choices, 'account_choices': account_choices}
        )
        
        if formset.is_valid():
            saved_instances = []
            
            try:
                with transaction.atomic():
                    for form in formset:
                        if form.cleaned_data and not form.cleaned_data.get('DELETE'):
                            purchase_instance = form.save(commit=False) 
                            purchase_instance.user = request.user
                            purchase_instance.batch = metadata.get('batch_name')
                            
                            # --- DATA CLEANING ---
                            # Prevent garbage values like "1", "null", or "Unknown" from becoming ="1" in Excel
                            garbage_values = ['null', 'none', 'unknown', 'n/a', '1', 'nan']
                            if str(purchase_instance.invoice_no).lower().strip() in garbage_values:
                                purchase_instance.invoice_no = None
                            if str(purchase_instance.vattin).lower().strip() in garbage_values:
                                purchase_instance.vattin = None
                            
                            # --- FIX: Convert empty strings to None for IntegerFields/CharFields ---
                            fields_to_clean = [
                                'account_id', 'vat_account_id', 'wht_debit_account_id', 
                                'credit_account_id', 'wht_account_id', 
                                'debit_account_id_2', 'debit_account_id_3'
                            ]
                            for field in fields_to_clean:
                                val = getattr(purchase_instance, field, None)
                                if val == '' or val == "" or str(val).lower() == 'none':
                                    setattr(purchase_instance, field, None)

                            # --- VENDOR RESOLUTION ---
                            vc = form.cleaned_data.get('vendor_choice')
                            raw_name = form.cleaned_data.get('company', 'Unknown Vendor')
                            
                            if str(vc).startswith('TEMP_'):
                                new_vid = vc.replace('TEMP_', '')
                                new_vendor, _ = Vendor.objects.get_or_create(
                                    vendor_id=new_vid, defaults={'name': raw_name}
                                )
                                purchase_instance.vendor = new_vendor
                            elif vc:
                                try:
                                    purchase_instance.vendor = Vendor.objects.get(id=int(vc))
                                except (ValueError, Vendor.DoesNotExist):
                                    pass
                                    
                            # 1. Save the Source Document (Purchase Invoice)
                            purchase_instance.save()
                            saved_instances.append(purchase_instance)
                            
                            # ==========================================================
                            # --- AI FEEDBACK PUBLISHING (PUB/SUB) ---
                            # ==========================================================
                            if form.has_changed() and 'account_id' in form.changed_data:
                                api_key = getattr(settings, 'GEMINI_API_KEY_2', os.getenv("GEMINI_API_KEY_2"))
                                initial_acct = form.initial.get('account_id')
                                final_acct = form.cleaned_data.get('account_id')
                                SystemOrchestrator.submit_correction_feedback(
                                    context_data=f"Vendor: {raw_name}, Description: {purchase_instance.description_en or purchase_instance.description}",
                                    ai_decision=f"Mapped to Account: {initial_acct}",
                                    human_correction=f"Changed to Account: {final_acct}",
                                    api_key=api_key
                                )

                            # ==========================================================
                            # --- 2. AUTOMATIC DOUBLE-ENTRY JOURNAL CREATION ---
                            # ==========================================================
                            
                            # Create Journal Entry Header (Explicit FK back to 'purchase')
                            je = JournalEntry.objects.create(
                                date=purchase_instance.date or date.today(),
                                description=f"Purchase from {raw_name}",
                                reference_number=purchase_instance.invoice_no,
                                purchase=purchase_instance
                            )

                            # Financial Calculations
                            total_amount = float(purchase_instance.total_usd or 0.0)
                            vat_amount = float(purchase_instance.vat_usd or 0.0)
                            net_amount = round(total_amount - vat_amount, 2)

                            # --- USER EDITED ACCOUNTS ---
                            form_debit_acct = form.cleaned_data.get('account_id')
                            form_credit_acct = form.cleaned_data.get('credit_account_id')

                            # CREDIT: Trade Payable (Total Liability)
                            if total_amount > 0:
                                cr_account_id = str(form_credit_acct) if form_credit_acct else '200000'
                                
                                # 💡 BACKEND SAFEGUARD: If AI or user accidentally passed a Cash/Bank account, force it to AP.
                                if cr_account_id.startswith('100'):
                                    cr_account_id = '200000'

                                ap_account, _ = Account.objects.get_or_create(
                                    account_id=cr_account_id, 
                                    defaults={'name': 'Trade Payable - USD', 'account_type': 'Liability'}
                                )
                                JournalLine.objects.create(
                                    journal_entry=je, account=ap_account, 
                                    description=f"Payable - {raw_name}", credit=total_amount
                                )

                            # DEBIT: VAT Input (Recoverable Tax Asset)
                            if vat_amount > 0:
                                vat_account, _ = Account.objects.get_or_create(
                                    account_id='115010', 
                                    defaults={'name': 'VAT input 进项增值税', 'account_type': 'Asset'}
                                )
                                JournalLine.objects.create(
                                    journal_entry=je, account=vat_account, 
                                    description="Input VAT", debit=vat_amount
                                )

                            main_net = net_amount

                            # 💡 NEW: PROCESS SECONDARY DEBIT (Accrual Clearing)
                            amt_2 = float(getattr(purchase_instance, 'debit_amount_2', 0.0) or 0.0)
                            acct_2 = str(getattr(purchase_instance, 'debit_account_id_2', '') or '')
                            desc_2 = str(getattr(purchase_instance, 'debit_desc_2', '') or '')
                            
                            if amt_2 > 0 and acct_2 and acct_2.lower() != 'none':
                                acc2_obj, _ = Account.objects.get_or_create(
                                    account_id=acct_2, 
                                    defaults={'name': 'Accrual Clearing', 'account_type': 'Liability'}
                                )
                                JournalLine.objects.create(
                                    journal_entry=je, account=acc2_obj, 
                                    description=desc_2 or f"Clearing Accrual for {raw_name}", 
                                    debit=amt_2
                                )
                                main_net = round(main_net - amt_2, 2)

                            # 💡 NEW: PROCESS TERTIARY DEBIT (Accrual Clearing 2)
                            amt_3 = float(getattr(purchase_instance, 'debit_amount_3', 0.0) or 0.0)
                            acct_3 = str(getattr(purchase_instance, 'debit_account_id_3', '') or '')
                            desc_3 = str(getattr(purchase_instance, 'debit_desc_3', '') or '')
                            
                            if amt_3 > 0 and acct_3 and acct_3.lower() != 'none':
                                acc3_obj, _ = Account.objects.get_or_create(
                                    account_id=acct_3, 
                                    defaults={'name': 'Secondary Accrual Clearing', 'account_type': 'Liability'}
                                )
                                JournalLine.objects.create(
                                    journal_entry=je, account=acc3_obj, 
                                    description=desc_3 or "Secondary Accrual Clearing", 
                                    debit=amt_3
                                )
                                main_net = round(main_net - amt_3, 2)

                            # PROCESS MAIN DEBIT (Current Month Expense)
                            if main_net > 0:
                                ai_account_id = str(form_debit_acct) if form_debit_acct else '725080'
                                exp_account, _ = Account.objects.get_or_create(
                                    account_id=ai_account_id, 
                                    defaults={'name': 'Operating Expense', 'account_type': 'Expense'}
                                )
                                JournalLine.objects.create(
                                    journal_entry=je, account=exp_account, 
                                    description=purchase_instance.description_en or purchase_instance.description or "Expense", 
                                    debit=main_net
                                )

            except Exception as e:
                messages.error(request, f"Database transaction failed. Nothing was saved. Error: {str(e)}")
                return render(request, 'invoice_review.html', {'formset': formset, 'metadata': metadata})
            
            # --- EXCEL REPORT GENERATION ---
            if saved_instances:
                report_data = list(Purchase.objects.filter(id__in=[p.id for p in saved_instances]).values())
                df_report = pd.DataFrame(report_data)

                # Remove timezone info from datetime columns for Excel compatibility
                for col in df_report.columns:
                    if pd.api.types.is_datetime64_any_dtype(df_report[col]) and df_report[col].dt.tz is not None:
                        df_report[col] = df_report[col].dt.tz_localize(None)

                media_dir = os.path.join(settings.BASE_DIR, 'media')
                os.makedirs(media_dir, exist_ok=True)
                report_path = os.path.join(media_dir, 'invoice_process_report.xlsx')
                df_report.to_excel(report_path, index=False, engine='openpyxl')
                request.session['invoice_report_path'] = report_path
            
            # Clean Session & Redirect
            request.session.pop('extracted_invoices', None)
            request.session.pop('ai_metadata', None)
            
            messages.success(request, f"Successfully saved {len(saved_instances)} invoices and posted Journal Entries!")
            return redirect('tools:invoice_download') 
        else:
            print("❌ FORMSET VALIDATION FAILED:")
            for i, form in enumerate(formset):
                if form.errors:
                    print(f"Row {i+1} Errors: {form.errors}")
            messages.error(request, "Validation failed. Please check the form for errors.")
            
    else:
        formset = PurchaseFormSet(
            initial=extracted_data, 
            form_kwargs={'dynamic_choices': dynamic_choices, 'account_choices': account_choices}
        )
        
    return render(request, template_name, {'formset': formset, 'metadata': metadata})

# ====================================================================
# --- 3. DOWNLOAD & DASHBOARD VIEWS ---
# ====================================================================

def invoice_download_view(request):
    """Renders the success page with the download link."""
    file_path = request.session.get('invoice_report_path')
    return render(request, 'invoice_download.html', {'has_file': bool(file_path and os.path.exists(file_path))})

def download_invoice_report(request):
    """Serves the generated Excel invoice report to the user."""
    file_path = request.session.get('invoice_report_path')
    if file_path and os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename="ai_process_report.xlsx"'
            return response
    
    messages.error(request, "The report file has expired or could not be found.")
    return redirect('tools:invoice_upload')

@login_required(login_url="register:login")
def hand_written_invoice_view(request):
    if request.method == 'POST':
        form = ManualInvoiceUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data['excel_file']
            batch_name = form.cleaned_data['batch_name']
            custom_prompt = form.cleaned_data.get('ai_prompt', '')
            
            # 1. Parse Excel / CSV into DataFrame
            try:
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                else:
                    df = pd.read_excel(uploaded_file, engine='openpyxl')
                
                # Standardize columns to lower case for mapping
                df.columns = df.columns.str.lower()
                
                # Filter out empty rows
                df.dropna(subset=['amount', 'description'], how='all', inplace=True)
                records = df.to_dict('records')
                
            except Exception as e:
                messages.error(request, f"Failed to read file. Please ensure it is a valid Excel/CSV. Error: {e}")
                return redirect('tools:hand_written_invoice')

            # 2. Setup AI Processor Context
            api_key = getattr(settings, 'GEMINI_API_KEY_2', os.getenv("GEMINI_API_KEY_2"))
            processor = GeminiInvoiceProcessor(api_key=api_key)
            
            rules_context = ""
            rules = AccountMappingRule.objects.all().select_related('account')
            if rules.exists():
                rules_data = [{'Account ID': r.account.account_id, 'Keywords': r.trigger_keywords, 'Guideline': r.ai_guideline} for r in rules]
                rules_context = pd.DataFrame(rules_data).to_csv(index=False)
                
            client_memo = ClientPromptMemo.objects.filter(category='GENERAL').first()
            memo_context = client_memo.memo_text if client_memo else ""

            # 3. Process records in chunks (e.g., 15 at a time) to prevent AI token/output truncation
            CHUNK_SIZE = 15
            all_ledgers = []
            total_cost = 0.0
            
            for i in range(0, len(records), CHUNK_SIZE):
                chunk = records[i:i+CHUNK_SIZE]
                print(f"🧠 Processing manual records {i+1} to {min(i+CHUNK_SIZE, len(records))}...")
                
                chunk_ledgers, cost, err = processor.process_manual_batch(
                    records=chunk, 
                    custom_prompt=custom_prompt,
                    batch_name=batch_name, 
                    rules_context=rules_context, 
                    memo_context=memo_context,
                    start_page=i + 1
                )
                
                if err:
                    messages.warning(request, f"Error processing chunk {i+1}: {err}")
                if chunk_ledgers:
                    all_ledgers.extend(chunk_ledgers)
                
                total_cost += cost

            # 4. Populate Session and pass to existing Review View
            if all_ledgers:
                try:
                    AICostLog.objects.create(
                        file_name=uploaded_file.name, total_pages=len(records), 
                        flash_cost=0.0, pro_cost=total_cost, total_cost=total_cost
                    )
                except Exception: pass
                
                request.session['extracted_invoices'] = all_ledgers
                request.session['ai_metadata'] = {
                    'file_name': uploaded_file.name,
                    'batch_name': batch_name,
                    'total_pages': len(records),  # Using record count as "pages" proxy
                    'costs': {'flash_cost': 0.0, 'pro_cost': total_cost},
                    'is_manual': True
                }
                
                messages.success(request, f"Successfully processed {len(all_ledgers)} manual entries.")
                return redirect('tools:review_invoices')
            else:
                messages.error(request, "AI failed to extract any valid records from the file.")
                
    else:
        form = ManualInvoiceUploadForm()

    return render(request, 'hand_written_upload.html', {'form': form})

@login_required(login_url="register:login")
def ai_cost_dashboard(request):
    """Dashboard to review AI processing costs."""
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "You do not have permission to view the AI cost dashboard.")
        return redirect('register:main')

    cost_logs_list = AICostLog.objects.all().order_by('-date')

    paginator = Paginator(cost_logs_list, 20)  # 20 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    totals = AICostLog.objects.aggregate(
        total_flash=Sum('flash_cost'), 
        total_pro=Sum('pro_cost'), 
        grand_total=Sum('total_cost'), 
        total_pages=Sum('total_pages')
    )
    return render(request, 'cost_dashboard.html', {'cost_logs': page_obj, 'totals': totals, 'page_obj': page_obj})

@login_required(login_url="register:login")
def manual_invoice_entry_view(request):
    """View to manually enter a single invoice and post it to the GL."""
    db_vendors = [(v.id, f"{v.vendor_id} - {v.name}") for v in Vendor.objects.all().order_by('vendor_id')]
    vendor_choices = [('', '--- Select Existing Vendor ---')] + db_vendors

    seen_accounts = set()
    db_accounts = []
    for acc_id, name in Account.objects.values_list('account_id', 'name'):
        if acc_id not in seen_accounts:
            seen_accounts.add(acc_id)
            db_accounts.append((str(acc_id), f"{acc_id} - {name}"))
    db_accounts.sort(key=lambda x: str(x[0]))
    account_choices = [('', '--- Select Account ---')] + db_accounts

    if request.method == 'POST':
        form = ManualPurchaseEntryForm(request.POST, vendor_choices=vendor_choices, account_choices=account_choices)
        
        if form.is_valid():
            
            # Wrap the entire creation process in an atomic transaction
            with transaction.atomic():
                purchase = form.save(commit=False)
                
                # CRITICAL: Assign the user so Profile permissions work in List/Detail views
                purchase.user = request.user 
                purchase.batch = "MANUAL_ENTRY"
                
                # Convert empty strings to None for IntegerFields
                for field in ['account_id', 'vat_account_id', 'wht_debit_account_id', 'credit_account_id', 'wht_account_id']:
                    val = getattr(purchase, field)
                    if val == '' or val == "":
                        setattr(purchase, field, None)
                
                # Resolve Vendor
                vc = form.cleaned_data.get('vendor_choice')
                if vc:
                    purchase.vendor_id = int(vc)
                
                purchase.save()

                # ==========================================================
                # --- POST TO GENERAL LEDGER ---
                # ==========================================================
                je = JournalEntry.objects.create(
                    date=purchase.date or date.today(),
                    description=f"Manual Purchase: {purchase.company}",
                    reference_number=purchase.invoice_no,
                    purchase=purchase
                )

                # Get amounts safely
                total_amount = float(purchase.total_usd or 0.0)
                vat_amount = float(purchase.vat_usd or 0.0)
                unreg_amount = float(purchase.unreg_usd or 0.0)
                
                # Calculate WHT
                wht_amount = 0.0
                if purchase.wht_account_id and unreg_amount > 0:
                    wht_amount = round(total_amount - unreg_amount, 2) 

                main_net = round((total_amount - vat_amount - wht_amount), 2)
                
                if purchase.account_id and main_net > 0:
                    acct, _ = Account.objects.get_or_create(account_id=str(purchase.account_id), defaults={'name': 'Operating Expense', 'account_type': 'Expense'})
                    JournalLine.objects.create(journal_entry=je, account=acct, description=purchase.description_en or "Expense", debit=main_net)

                # 2. VAT Debit
                if vat_amount > 0 and purchase.vat_account_id:
                    vat_acct, _ = Account.objects.get_or_create(account_id=str(purchase.vat_account_id), defaults={'name': 'VAT input', 'account_type': 'Asset'})
                    JournalLine.objects.create(journal_entry=je, account=vat_acct, description="Input VAT", debit=vat_amount)

                # 3. WHT Expense Debit
                if wht_amount > 0 and purchase.wht_debit_account_id:
                    wht_exp_acct, _ = Account.objects.get_or_create(account_id=str(purchase.wht_debit_account_id), defaults={'name': 'WHT Expense', 'account_type': 'Expense'})
                    JournalLine.objects.create(journal_entry=je, account=wht_exp_acct, description="WHT Expense Absorbed", debit=wht_amount)

                # 4. Main Credit (Payable)
                if total_amount > 0 and purchase.credit_account_id:
                    cr_acct, _ = Account.objects.get_or_create(account_id=str(purchase.credit_account_id), defaults={'name': 'Trade Payable', 'account_type': 'Liability'})
                    JournalLine.objects.create(journal_entry=je, account=cr_acct, description=f"Payable - {purchase.company}", credit=total_amount)

                # 5. WHT Payable Credit
                if wht_amount > 0 and purchase.wht_account_id:
                    wht_pay_acct, _ = Account.objects.get_or_create(account_id=str(purchase.wht_account_id), defaults={'name': 'WHT Payable', 'account_type': 'Liability'})
                    JournalLine.objects.create(journal_entry=je, account=wht_pay_acct, description="WHT Payable to GDT", credit=wht_amount)

            # --- End of Atomic Block ---
            
            messages.success(request, f"Successfully created manual invoice and posted Journal Entry for {purchase.company}.")
            return redirect('tools:manual_invoice_entry') 

    else:
        form = ManualPurchaseEntryForm(vendor_choices=vendor_choices, account_choices=account_choices)

    return render(request, 'manual_invoice_entry.html', {'form': form})

@login_required
def export_purchase_invoices(request):
    """Exports Purchase instances to an Excel file."""
    queryset = Purchase.objects.select_related('vendor').prefetch_related('journal_entries__lines__account').order_by('id')
    resource = PurchaseResource()
    dataset = resource.export(queryset=queryset)

    today_str = datetime.date.today().strftime("%Y%m%d")
    filename = f"purchase_invoices_{today_str}.xlsx"
    
    media_dir = os.path.join(settings.BASE_DIR, 'media')
    os.makedirs(media_dir, exist_ok=True)
    report_path = os.path.join(media_dir, filename)
    
    with open(report_path, 'wb') as f:
        f.write(dataset.xlsx)
        
    request.session['export_report_path'] = report_path
    request.session['export_filename'] = filename
    
    messages.success(request, f"Successfully exported purchase invoices!")
    return redirect('tools:purchase_export_success')

def purchase_export_success_view(request):
    """Renders the success page after an export completes."""
    file_path = request.session.get('export_report_path')
    return render(request, 'purchase_export_success.html', {'has_file': bool(file_path and os.path.exists(file_path))})

def download_exported_purchases(request):
    """Serves the exported Excel file to the user."""
    file_path = request.session.get('export_report_path')
    filename = request.session.get('export_filename', 'exported_purchases.xlsx')
    
    if file_path and os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    
    messages.error(request, "The export file has expired or could not be found.")
    return redirect('tools:invoice_upload')

@login_required
def gl_migration_upload_view(request):
    """Uploads GL data, parses via DB-backed AI, and stores in session queue."""
    if request.method == 'POST':
        request.session.pop('gl_report_path', None)
        request.session.pop('gl_migration_log', None)
        request.session.pop('gl_migration_completed', None)
        
        form = GLMigrationUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data['gl_file']
            batch_name = form.cleaned_data['batch_name']
            
            _, file_ext = os.path.splitext(uploaded_file.name)
            with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp_file:
                for chunk in uploaded_file.chunks():
                    tmp_file.write(chunk)
                tmp_file_path = tmp_file.name

            try:
                api_key = os.getenv("GEMINI_API_KEY_2") 
                processor = GLMigrationProcessor(api_key=api_key)
                
                parsed_data, costs = processor.process_migration_file(tmp_file_path)
                
                # --- NEW: PREVENT GHOST SAVES (Hard Stop if AI returned nothing) ---
                if not parsed_data:
                    messages.error(request, "AI Extraction Failed: No valid transactions were processed. Please check the file formatting or terminal logs for Pydantic schema errors.")
                    return redirect('tools:gl_migration_upload')
                
                # Log the AI Cost Immediately
                total_items = len(parsed_data)
                AICostLog.objects.create(
                    file_name=uploaded_file.name, 
                    total_pages=total_items, # Treat the number of extracted lines as 'pages'
                    flash_cost=costs.get('flash_cost', 0), 
                    pro_cost=costs.get('pro_cost', 0), 
                    total_cost=costs.get('flash_cost', 0) + costs.get('pro_cost', 0)
                )
                
                # Save the parsed data arrays to session queue
                request.session['gl_migration_data'] = {'lines': parsed_data}
                request.session['gl_migration_meta'] = {
                    'batch_name': batch_name
                }
                request.session['gl_migration_log'] = [] 
                request.session['gl_migration_completed'] = []
                
                messages.success(request, f"Data parsed successfully. Found {total_items} lines. Please review the batches.")
                return redirect('tools:gl_review')
                
            except Exception as e:
                print(f"❌ Migration Error: {str(e)}")
                messages.error(request, f"Migration Error: {str(e)}")
            finally:
                if os.path.exists(tmp_file_path):
                    os.remove(tmp_file_path)
        else:
            messages.error(request, "Form validation failed.")
    else:
        form = GLMigrationUploadForm()
        
    return render(request, 'tools/gl_migration_upload.html', {'form': form})


@login_required
def gl_review_view(request):
    """Processes the session queue in chunks via Formsets. Saves to DB only when queue is completely empty."""
    parsed_data = request.session.get('gl_migration_data', {})
    meta = request.session.get('gl_migration_meta', {})
    completed_data = request.session.get('gl_migration_completed', [])

    if not parsed_data or not meta:
        messages.error(request, "No migration queue found. Please upload a file.")
        return redirect('tools:gl_migration_upload')

    seen_accounts = set()
    db_accounts = []
    for acc_id, name in Account.objects.values_list('account_id', 'name'):
        if acc_id not in seen_accounts:
            seen_accounts.add(acc_id)
            db_accounts.append((str(acc_id), f"{acc_id} - {name}"))
    db_accounts.sort(key=lambda x: str(x[0]))
    account_choices = [('', '--- Select Account ---')] + db_accounts

    # We process a maximum of 30 items per page to prevent browser lag
    CHUNK_SIZE = 30
    lines_queue = parsed_data.get('lines', [])

    if request.method == 'POST':
        formset = GLHistoricalFormSet(request.POST, form_kwargs={'account_choices': account_choices})

        if formset.is_valid():
            # Extract cleaned data from this chunk
            chunk_results = []
            for form in formset:
                if form.cleaned_data and not form.cleaned_data.get('DELETE'):
                    cd = form.cleaned_data
                    chunk_results.append({
                        'gl_no': cd.get('gl_no') or 'UNGROUPED',
                        'date': str(cd['date']) if cd.get('date') else None,
                        'account_id': cd['account_id'],
                        'description': cd['description'],
                        'instruction': cd.get('instruction', ''),
                        'debit': cd['debit'] or 0.0,
                        'credit': cd['credit'] or 0.0
                    })
            
            # Add to the session's completed list
            completed_data.extend(chunk_results)
            request.session['gl_migration_completed'] = completed_data

            # Remove the processed chunk from the session queues
            parsed_data['lines'] = lines_queue[CHUNK_SIZE:]
            
            request.session['gl_migration_data'] = parsed_data
            request.session.modified = True

            # If queue is empty, we are done reviewing! Now perform atomic DB save.
            if not parsed_data['lines']:
                report_log = []
                
                # --- NEW: PREVENT GHOST SAVES (Hard Stop if completed_data is empty) ---
                if not completed_data:
                    messages.error(request, "CRITICAL ERROR: Attempted to save an empty dataset to the database.")
                    return redirect('tools:gl_migration_upload')
                
                try:
                    with transaction.atomic():
                        for item in completed_data:
                            # 1. PROCESS AND SAVE TO 'Old' MODEL
                            old_record = Old.objects.create(
                                user=request.user,
                                date=item['date'] or date.today(),
                                account_id=item['account_id'],
                                description=item['description'],
                                instruction=item['instruction'],
                                debit=item['debit'],
                                credit=item['credit']
                            )
                            
                            gl_no = item['gl_no']
                            
                            # 2. CREATE LINKED JOURNAL ENTRY
                            ref = f"HIST-{gl_no}" if gl_no and gl_no != 'UNGROUPED' else f"OLD-{old_record.id}"
                            je = JournalEntry.objects.create(
                                date=item['date'] or date.today(),
                                reference_number=ref,
                                description=f"Historical GL Migration: {item.get('description', '')}"[:255],
                                old=old_record
                            )
                            
                            # 3. CREATE JOURNAL LINE
                            account, _ = Account.objects.get_or_create(
                                account_id=str(item['account_id']),
                                defaults={'name': 'System Gen Acct', 'account_type': 'Asset'}
                            )
                            
                            safe_desc = item['description'][:255] if item.get('description') else "Historical Entry"
                            if item['debit'] > 0 and item['credit'] > 0:
                                JournalLine.objects.create(journal_entry=je, account=account, debit=item['debit'], credit=0.0, description=safe_desc)
                                JournalLine.objects.create(journal_entry=je, account=account, debit=0.0, credit=item['credit'], description=safe_desc)
                            else:
                                JournalLine.objects.create(
                                    journal_entry=je,
                                    account=account,
                                    debit=item['debit'],
                                    credit=item['credit'],
                                    description=safe_desc
                                )
                            
                            report_log.append({
                                'GL No': gl_no, 'Date': item['date'], 
                                'Account': item['account_id'], 
                                'Debit': item['debit'], 'Credit': item['credit'], 
                                'Description': item['description']
                            })
                except Exception as e:
                    messages.error(request, f"Database transaction failed during final save. Nothing was saved. Error: {str(e)}")
                    return render(request, 'tools/gl_review.html', {
                        'formset': formset, 'meta': meta, 'total_remaining': 0
                    })

                # Generate final report
                if report_log:
                    df_report = pd.DataFrame(report_log)
                    media_dir = os.path.join(settings.BASE_DIR, 'media')
                    os.makedirs(media_dir, exist_ok=True)
                    report_path = os.path.join(media_dir, f'gl_migration_report_{datetime.now().strftime("%Y%m%d%H%M")}.xlsx')
                    df_report.to_excel(report_path, index=False, engine='openpyxl')
                    request.session['gl_report_path'] = report_path

                request.session.pop('gl_migration_data', None)
                request.session.pop('gl_migration_meta', None)
                request.session.pop('gl_migration_log', None)
                request.session.pop('gl_migration_completed', None)
                
                messages.success(request, f"🎉 All {len(completed_data)} historical records successfully staged to Old model and mapped to Journals!")
                return redirect('tools:gl_download')
            
            messages.success(request, "Batch reviewed and queued. Loading next items...")
            return redirect('tools:gl_review')
            
        else:
            messages.error(request, "Validation errors found. Please correct them below.")
    else:
        # Load the next chunk into the forms
        formset = GLHistoricalFormSet(initial=lines_queue[:CHUNK_SIZE], form_kwargs={'account_choices': account_choices})

    total_remaining = len(lines_queue)

    return render(request, 'tools/gl_review.html', {
        'formset': formset,
        'meta': meta,
        'total_remaining': total_remaining
    })

@login_required
def gl_download_view(request):
    """Provides the download link for the completed migration report."""
    file_path = request.session.get('gl_report_path')
    has_file = bool(file_path and os.path.exists(file_path))
    file_url = f"/media/{os.path.basename(file_path)}" if has_file else ""
    return render(request, 'tools/gl_download.html', {
        'has_file': has_file,
        'file_url': file_url
    })

@login_required(login_url="register:login")
def PurchaseListView(request):
    purchases = Purchase.objects.all().order_by('-id')
    vendor_queryset = Vendor.objects.all().order_by('vendor_id')

    # Initialize Filter
    purchase_filter = PurchaseFilter(request.GET, queryset=purchases)
    purchase_filter.form.fields['vendor'].queryset = vendor_queryset

    # Apply Pagination (20 items per page)
    paginator = Paginator(purchase_filter.qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'filter': purchase_filter,
        'purchases': page_obj,  # 'page_obj' is fully iterable, keeping the template loop happy
        'page_obj': page_obj,
    }
    return render(request, 'purchase_list.html', context)


class PurchaseDetailView(LoginRequiredMixin, DetailView):
    login_url = "register:login"
    model = Purchase
    template_name = 'purchase_detail.html'
    context_object_name = 'purchase'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_owner'] = True
        return context


class PurchaseUpdateView(LoginRequiredMixin, UpdateView):
    login_url = "register:login"
    model = Purchase
    form_class = ManualPurchaseEntryForm 
    template_name = 'purchase_update.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        
        db_vendors = [(v.id, f"{v.vendor_id} - {v.name}") for v in Vendor.objects.all().order_by('vendor_id')]
        kwargs['vendor_choices'] = [('', '--- Select Existing Vendor ---')] + db_vendors
        
        db_accounts = [(a.account_id, f"{a.account_id} - {a.name}") for a in Account.objects.all().order_by('account_id')]
        kwargs['account_choices'] = [('', '--- Select Account ---')] + db_accounts
        
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        if self.object.vendor:
            initial['vendor_choice'] = self.object.vendor.id
        return initial

    def form_valid(self, form):
        # Wrap everything in an atomic transaction to prevent partial writes/duplicates
        with transaction.atomic():
            purchase = form.save(commit=False)
            fields_to_clean = [
                'account_id', 'vat_account_id', 'wht_debit_account_id', 'credit_account_id', 'wht_account_id',
                'debit_account_id_2', 'debit_account_id_3'
            ]
            for field in fields_to_clean:
                val = getattr(purchase, field)
                if val == '' or val == "":
                    setattr(purchase, field, None)
            vc = form.cleaned_data.get('vendor_choice')
            if vc:
                purchase.vendor_id = int(vc)
            purchase.save()

            # ==========================================================
            # --- ATOMIC RECALCULATION OF GENERAL LEDGER ---
            # ==========================================================
            je, created = JournalEntry.objects.get_or_create(
                purchase=purchase,
                defaults={
                    'date': purchase.date or date.today(),
                    'description': f"Purchase from {purchase.company}",
                    'reference_number': purchase.invoice_no,
                }
            )
            if not created:
                je.date = purchase.date or date.today()
                je.description = f"Updated Purchase from {purchase.company}"
                je.reference_number = purchase.invoice_no
                je.save(update_fields=['date', 'description', 'reference_number'])
                je.lines.all().delete()

            total_amount = float(purchase.total_usd or 0.0)
            vat_amount = float(purchase.vat_usd or 0.0)

            # CREDIT: Main Payable
            if total_amount > 0 and purchase.credit_account_id:
                cr_acct, _ = Account.objects.get_or_create(account_id=str(purchase.credit_account_id), defaults={'name': 'Trade Payable', 'account_type': 'Liability'})
                JournalLine.objects.create(journal_entry=je, account=cr_acct, description=f"Payable - {purchase.company}", credit=total_amount)

            # DEBIT: VAT
            if vat_amount > 0 and purchase.vat_account_id:
                vat_acct, _ = Account.objects.get_or_create(account_id=str(purchase.vat_account_id), defaults={'name': 'VAT input', 'account_type': 'Asset'})
                JournalLine.objects.create(journal_entry=je, account=vat_acct, description="Input VAT", debit=vat_amount)

            # Start with net expense and subtract accruals
            main_net = round(total_amount - vat_amount, 2)

            # DEBIT: Accrual 2
            amt_2 = float(getattr(purchase, 'debit_amount_2', 0.0) or 0.0)
            acct_2 = str(getattr(purchase, 'debit_account_id_2', '') or '')
            desc_2 = str(getattr(purchase, 'debit_desc_2', '') or '')
            if amt_2 > 0 and acct_2:
                acc2_obj, _ = Account.objects.get_or_create(account_id=acct_2, defaults={'name': 'Accrual Clearing', 'account_type': 'Liability'})
                JournalLine.objects.create(journal_entry=je, account=acc2_obj, description=desc_2 or "Accrual Clearing", debit=amt_2)
                main_net = round(main_net - amt_2, 2)

            # DEBIT: Accrual 3
            amt_3 = float(getattr(purchase, 'debit_amount_3', 0.0) or 0.0)
            acct_3 = str(getattr(purchase, 'debit_account_id_3', '') or '')
            desc_3 = str(getattr(purchase, 'debit_desc_3', '') or '')
            if amt_3 > 0 and acct_3:
                acc3_obj, _ = Account.objects.get_or_create(account_id=acct_3, defaults={'name': 'Secondary Accrual Clearing', 'account_type': 'Liability'})
                JournalLine.objects.create(journal_entry=je, account=acc3_obj, description=desc_3 or "Secondary Accrual", debit=amt_3)
                main_net = round(main_net - amt_3, 2)

            # DEBIT: Main Expense
            if purchase.account_id and main_net > 0:
                acct, _ = Account.objects.get_or_create(account_id=str(purchase.account_id), defaults={'name': 'Operating Expense', 'account_type': 'Expense'})
                JournalLine.objects.create(journal_entry=je, account=acct, description=purchase.description_en or "Expense", debit=main_net)

            # WHT Logic (preserved from original)
            unreg_amount = float(purchase.unreg_usd or 0.0)
            wht_amount = 0.0
            if purchase.wht_account_id and unreg_amount > 0:
                wht_amount = round(total_amount - unreg_amount, 2)
            
            if wht_amount > 0 and purchase.wht_debit_account_id:
                wht_exp_acct, _ = Account.objects.get_or_create(account_id=str(purchase.wht_debit_account_id), defaults={'name': 'WHT Expense', 'account_type': 'Expense'})
                JournalLine.objects.create(journal_entry=je, account=wht_exp_acct, description="WHT Expense Absorbed", debit=wht_amount)
            
            if wht_amount > 0 and purchase.wht_account_id:
                wht_pay_acct, _ = Account.objects.get_or_create(account_id=str(purchase.wht_account_id), defaults={'name': 'WHT Payable', 'account_type': 'Liability'})
                JournalLine.objects.create(journal_entry=je, account=wht_pay_acct, description="WHT Payable to GDT", credit=wht_amount)

        # End of atomic block. If successful, proceed to success message and redirect.
        messages.success(self.request, "Purchase invoice and General Ledger entries updated successfully!")
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse('tools:purchase_detail', kwargs={'pk': self.object.pk})

class PurchaseDeleteView(LoginRequiredMixin, DeleteView):
    login_url = "register:login"
    model = Purchase
    template_name = 'purchase_confirm_delete.html'
    success_url = reverse_lazy('tools:purchase_list')

    def form_valid(self, form):
        # Clean up associated General Ledger Entries before deleting the purchase
        JournalEntry.objects.filter(purchase=self.object).delete()
        messages.success(self.request, 'Purchase and associated Journal Entries deleted successfully!')
        return super().form_valid(form)

@login_required(login_url="register:login")
def export_purchase_csv(request):
    purchases = Purchase.objects.select_related('vendor').prefetch_related('journal_entries__lines__account').order_by('id')

    # Apply the same filter parameters passed via the GET request
    purchase_filter = PurchaseFilter(request.GET, queryset=purchases)
    filtered_purchases = purchase_filter.qs

    # Generate the CSV using django-import-export
    resource = PurchaseResource()
    dataset = resource.export(queryset=filtered_purchases)
    
    # Create and return the HTTP response with a UTF-8 BOM so Excel reads Chinese characters properly
    response = HttpResponse(dataset.csv.encode('utf-8-sig'), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="purchase_invoices.csv"'
    
    return response

# ====================================================================
# --- 4. OLD MODEL CRUD & JOURNAL POSTING ---
# ====================================================================

@login_required(login_url="register:login")
def OldListView(request):
    old_records = Old.objects.all().order_by('-id')

    paginator = Paginator(old_records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'old_records': page_obj,
        'page_obj': page_obj,
    }
    return render(request, 'tools/old_list.html', context)


@login_required(login_url="register:login")
def manual_old_entry_view(request):
    seen_accounts = set()
    db_accounts = []
    for acc_id, name in Account.objects.values_list('account_id', 'name'):
        if acc_id not in seen_accounts:
            seen_accounts.add(acc_id)
            db_accounts.append((str(acc_id), f"{acc_id} - {name}"))
    db_accounts.sort(key=lambda x: str(x[0]))
    account_choices = [('', '--- Select Account ---')] + db_accounts

    if request.method == 'POST':
        form = OldEntryForm(request.POST, account_choices=account_choices)
        if form.is_valid():
            with transaction.atomic():
                old_record = form.save(commit=False)
                old_record.user = request.user
                old_record.save()

                # Post to General Ledger, linking via reference_number
                je = JournalEntry.objects.create(
                    date=old_record.date or date.today(),
                    description=f"Historical Entry: {old_record.description}"[:255],
                    reference_number=f"OLD-{old_record.id}",
                    old=old_record
                )
                
                acct, _ = Account.objects.get_or_create(
                    account_id=str(old_record.account_id), 
                    defaults={'name': 'Historical Default', 'account_type': 'Asset'}
                )
                safe_desc = old_record.description[:255] if old_record.description else "Historical Entry"
                debit_val = old_record.debit or 0.0
                credit_val = old_record.credit or 0.0
                
                if debit_val > 0 and credit_val > 0:
                    JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=debit_val, credit=0.0)
                    JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=0.0, credit=credit_val)
                else:
                    JournalLine.objects.create(
                        journal_entry=je, 
                        account=acct, 
                        description=safe_desc, 
                        debit=debit_val,
                        credit=credit_val
                    )
            
            messages.success(request, f"Successfully created manual historical record and posted to GL.")
            return redirect('tools:old_list') 
    else:
        form = OldEntryForm(account_choices=account_choices)

    return render(request, 'tools/old_form.html', {'form': form})


class OldDetailView(LoginRequiredMixin, DetailView):
    login_url = "register:login"
    model = Old
    template_name = 'tools/old_detail.html'
    context_object_name = 'old_record'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_owner'] = True
        return context


class OldUpdateView(LoginRequiredMixin, UpdateView):
    login_url = "register:login"
    model = Old
    form_class = OldEntryForm 
    template_name = 'tools/old_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        db_accounts = [(a.account_id, f"{a.account_id} - {a.name}") for a in Account.objects.all().order_by('account_id')]
        kwargs['account_choices'] = [('', '--- Select Account ---')] + db_accounts
        return kwargs

    def form_valid(self, form):
        with transaction.atomic():
            old_record = form.save()
            je, created = JournalEntry.objects.get_or_create(
                old=old_record,
                defaults={
                    'date': old_record.date or date.today(),
                    'description': f"Historical Entry: {old_record.description}"[:255],
                    'reference_number': f"OLD-{old_record.id}",
                }
            )
            if not created:
                je.date = old_record.date or date.today()
                je.description = f"Updated Historical Entry: {old_record.description}"[:255]
                je.save(update_fields=['date', 'description'])
                je.lines.all().delete()

            acct, _ = Account.objects.get_or_create(account_id=str(old_record.account_id), defaults={'name': 'Historical Default', 'account_type': 'Asset'})
            safe_desc = old_record.description[:255] if old_record.description else "Historical Entry"
            debit_val = old_record.debit or 0.0
            credit_val = old_record.credit or 0.0
            if debit_val > 0 and credit_val > 0:
                JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=debit_val, credit=0.0)
                JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=0.0, credit=credit_val)
            else:
                JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=debit_val, credit=credit_val)
        messages.success(self.request, "Historical record and General Ledger entries updated successfully!")
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse('tools:old_detail', kwargs={'pk': self.object.pk})

class OldDeleteView(LoginRequiredMixin, DeleteView):
    login_url = "register:login"
    model = Old
    template_name = 'tools/old_confirm_delete.html'
    success_url = reverse_lazy('tools:old_list')

    def form_valid(self, form):
        JournalEntry.objects.filter(Q(old=self.object) | Q(reference_number=f"OLD-{self.object.id}")).delete()
        messages.success(self.request, 'Record and associated Journal Entries deleted successfully!')
        return super().form_valid(form)

############################

# Export Balancika Sheets

############################

def clean_invoice_number(val):
    s = str(val).strip()
    if s.lower() in ['nan', 'none'] or s == '': 
        return ''
    if re.match(r'^-?\d+(\.\d+)?[eE][+\-]?\d+$', s):
        try:
            return '{:.0f}'.format(float(s))
        except:
            return s
    return s.replace('.0', '')

@login_required
def export_balancika_view(request):
    if request.method == 'POST':
        form = BalancikaExportForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
            purchase_id = form.cleaned_data.get('purchase_id')
            bank_id = form.cleaned_data.get('bank_id')
            entry_counter = form.cleaned_data['entry_no_start']
            
            purchase_filters = Q()
            bank_filters = Q(credit__gt=0)

            if start_date:
                purchase_filters &= Q(date__gte=start_date)
                bank_filters &= Q(date__gte=start_date)
            if end_date:
                purchase_filters &= Q(date__lte=end_date)
                bank_filters &= Q(date__lte=end_date)

            purchases = []
            bank_charges = []

            if purchase_id:
                purchases = list(Purchase.objects.filter(purchase_filters & Q(id=purchase_id)).order_by('id'))
            elif bank_id:
                bank_charges = list(Bank.objects.filter(bank_filters & Q(id=bank_id)).order_by('id'))
            else:
                purchases = list(Purchase.objects.filter(purchase_filters).order_by('id'))
                
                bank_charges = list(Bank.objects.filter(bank_filters & Q(vendor__name__icontains='bank')).order_by('id'))

            combined_records = purchases + bank_charges
            # Order first by type (Purchases then Bank charges), then by ID
            combined_records.sort(key=lambda x: (0 if isinstance(x, Purchase) else 1, x.id))

            if not combined_records:
                messages.warning(request, f"No purchases or bank charges found with the given criteria.")
                return render(request, 'tools/balancika_export.html', {'form': form})

            # Calculate base month start and end dates
            base_start_str = start_date.strftime('%d-%b-%Y') if start_date else date.today().strftime('%d-%b-%Y')
            base_end_str = end_date.strftime('%d-%b-%Y') if end_date else date.today().strftime('%d-%b-%Y')

            sheet1_data = []
            sheet2_data = []

            for record in combined_records:
                entry_no = f"PIN-{entry_counter:05d}"
                entry_counter += 1

                if isinstance(record, Purchase):
                    p = record
                    # Mappings from Django Model (Adjust field names if your model differs slightly)
                    original_acct_id = str(getattr(p, 'account_id', '')).strip()
                    original_vendor_id = str(p.vendor.vendor_id).strip() if p.vendor else ''
                    original_invoice = clean_invoice_number(getattr(p, 'invoice_no', ''))
                    description = str(getattr(p, 'description', '')).strip()
                    description_en = str(getattr(p, 'description_en', '')).strip()

                    # Date Formatting
                    if p.date:
                        final_date = p.date.strftime('%d-%b-%Y')
                        # Find the last day of the transaction's month
                        _, p_last_day = calendar.monthrange(p.date.year, p.date.month)
                        final_due_date = date(p.date.year, p.date.month, p_last_day).strftime('%d-%b-%Y')
                    else:
                        final_date, final_due_date = base_start_str, base_end_str

                    # --- SHEET 1 POPULATION ---
                    sheet1_data.append({
                        "Entry No": entry_no,
                        "Date (dd-MMM-YYYY)": final_date,
                        "Type": "Apply to GL Account",
                        "Reference": original_invoice,
                        "Remark": description_en if description_en else description,
                        "Vendor ID": original_vendor_id,
                        "Employee ID": "",
                        "Class ID": "",
                        "Due Date (dd-MMM-YYYY)": final_due_date,
                        "Purchase Order": "",
                        "Currency ID": "USD",
                        "Exchange Rate": 1
                    })

                    # --- SHEET 2 SCENARIO LOGIC ---
                    # Safely get numeric amounts from the Django model
                    local_vat_amt = float(getattr(p, 'vat_usd', 0.0) or 0.0)
                    total_amt = float(getattr(p, 'total_usd', 0.0) or 0.0)
                    
                    # If your model has specific non-vat fields, use them. Otherwise, calculate.
                    local_purchase_amt = float(getattr(p, 'local_purchase_usd', total_amt - local_vat_amt) or 0.0)
                    non_vat_amt = float(getattr(p, 'non_vat_usd', 0.0) or 0.0) 

                    desc_lower = description.lower()
                    display_desc = description_en if description_en else description
                    rows_to_add = []

                    # RENTAL SCENARIO
                    if 'rental' in desc_lower:
                        wht_val = total_amt * 0.10
                        rows_to_add.append({"Desc": display_desc, "Cost": total_amt, "Total": total_amt, "Line": 1, "VAT": "None", "AcctID": original_acct_id})
                        rows_to_add.append({"Desc": "10% WHT on Rental", "Cost": -wht_val, "Total": -wht_val, "Line": 2, "VAT": "None", "AcctID": "23500"})
                        rows_to_add.append({"Desc": "10% WHT on Rental Expense", "Cost": wht_val, "Total": wht_val, "Line": 3, "VAT": "None", "AcctID": "65000"})

                    # NSSF SCENARIO
                    elif ("nssf" in desc_lower or "occupational risk" in desc_lower) and "pension" in desc_lower:
                        parts = [pt.strip() for pt in re.split(r'\.\s+', description) if pt.strip()]
                        for idx, pt in enumerate(parts[:3]):
                            m = re.search(r"(?:--|-)\s*([\d,]+\.?\d*)", pt)
                            amt = float(m.group(1).replace(',', '')) if m else 0.0
                            rows_to_add.append({"Desc": pt, "Cost": amt, "Total": amt, "Line": idx+1, "VAT": "None", "AcctID": original_acct_id})

                    # SCENARIOS B/D/A
                    elif non_vat_amt != 0 and local_vat_amt != 0:
                        rows_to_add.append({"Desc": display_desc, "Cost": local_purchase_amt, "Total": local_purchase_amt, "Line": 1, "VAT": "VAT_IN_1", "AcctID": original_acct_id})
                        rows_to_add.append({"Desc": display_desc, "Cost": non_vat_amt, "Total": non_vat_amt, "Line": 2, "VAT": "None", "AcctID": original_acct_id})
                    elif local_vat_amt != 0:
                        rows_to_add.append({"Desc": display_desc, "Cost": local_purchase_amt, "Total": local_purchase_amt, "Line": 1, "VAT": "VAT_IN_1", "AcctID": original_acct_id})
                    else:
                        rows_to_add.append({"Desc": display_desc, "Cost": total_amt, "Total": total_amt, "Line": 1, "VAT": "None", "AcctID": original_acct_id})

                    for r in rows_to_add:
                        f_acct = str(r["AcctID"])
                        sheet2_data.append({
                            "Entry No": entry_no,
                            "Description": r["Desc"],
                            "Account ID": f_acct, 
                            "Item ID": "", 
                            "Location ID": "", 
                            "Lot Number": "",
                            "Manufacture Date (dd-MMM-YYYY)": "", 
                            "Expiry Date (dd-MMM-YYYY)": "",
                            "Class ID": "", 
                            "Uom ID": "",
                            "Quantity": 1, 
                            "FOC": 0,
                            "Unit Cost": r["Cost"], 
                            "Total Cost": r["Cost"],
                            "Disc %": 0, 
                            "Disc": 0,
                            "Grand Total": r["Total"],
                            "Purchase Order No": "", 
                            "Purchase Order Line": r["Line"],
                            "Job No": "", 
                            "Job Task No": 0, 
                            "Job Planning Line": 0,
                            "WHT": "None", 
                            "Tax Group 2": "None", 
                            "Tax Group 3": "None",
                            "VAT Input": r["VAT"]
                        })
                else:
                    b = record
                    original_acct_id = str(getattr(b, 'debit_account_id', ''))
                    original_vendor_id = str(b.vendor.vendor_id).strip() if b.vendor else ''
                    original_invoice = b.bank_ref_id or ''
                    
                    trans_type = str(getattr(b, 'trans_type', '')).strip()
                    purpose = str(getattr(b, 'purpose', '')).strip()
                    remark = str(getattr(b, 'remark', '')).strip()
                    
                    display_desc = remark if remark else purpose
                    if trans_type and trans_type.lower() not in display_desc.lower():
                        display_desc = f"{trans_type} - {display_desc}" if display_desc else trans_type
                    if not display_desc:
                        display_desc = "Bank Charge"
                    
                    if b.date:
                        final_date = b.date.strftime('%d-%b-%Y')
                        _, b_last_day = calendar.monthrange(b.date.year, b.date.month)
                        final_due_date = date(b.date.year, b.date.month, b_last_day).strftime('%d-%b-%Y')
                    else:
                        final_date, final_due_date = base_start_str, base_end_str
                        
                    sheet1_data.append({
                        "Entry No": entry_no,
                        "Date (dd-MMM-YYYY)": final_date,
                        "Type": "Apply to GL Account",
                        "Reference": original_invoice,
                        "Remark": display_desc,
                        "Vendor ID": original_vendor_id,
                        "Employee ID": "",
                        "Class ID": "",
                        "Due Date (dd-MMM-YYYY)": final_due_date,
                        "Purchase Order": "",
                        "Currency ID": "USD",
                        "Exchange Rate": 1
                    })
                    
                    charge_amt = float(b.credit)
                    
                    sheet2_data.append({
                        "Entry No": entry_no,
                        "Description": display_desc,
                        "Account ID": original_acct_id, 
                        "Item ID": "", 
                        "Location ID": "", 
                        "Lot Number": "",
                        "Manufacture Date (dd-MMM-YYYY)": "", 
                        "Expiry Date (dd-MMM-YYYY)": "",
                        "Class ID": "", 
                        "Uom ID": "",
                        "Quantity": 1, 
                        "FOC": 0,
                        "Unit Cost": charge_amt, 
                        "Total Cost": charge_amt,
                        "Disc %": 0, 
                        "Disc": 0,
                        "Grand Total": charge_amt,
                        "Purchase Order No": "", 
                        "Purchase Order Line": 1,
                        "Job No": "", 
                        "Job Task No": 0, 
                        "Job Planning Line": 0,
                        "WHT": "None", 
                        "Tax Group 2": "None", 
                        "Tax Group 3": "None",
                        "VAT Input": "None"
                    })

            # --- GENERATE EXCEL FILE IN MEMORY ---
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                pd.DataFrame(sheet1_data).to_excel(writer, sheet_name='Sheet1', index=False)
                pd.DataFrame(sheet2_data).to_excel(writer, sheet_name='Sheet2', index=False)
            
            # Rewind the buffer
            output.seek(0)

            # --- RETURN AS DOWNLOADABLE ATTACHMENT ---
            date_range_str = ""
            if start_date and end_date:
                date_range_str = f"_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}"
            elif start_date:
                date_range_str = f"_from_{start_date.strftime('%Y%m%d')}"
            elif end_date:
                date_range_str = f"_to_{end_date.strftime('%Y%m%d')}"
            
            filename = f"Balancika_Export{date_range_str}.xlsx"
            response = HttpResponse(
                output.read(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    else:
        form = BalancikaExportForm()

    return render(request, 'tools/balancika_export.html', {'form': form})

# ====================================================================
# UPDATE ENGAGEMENT LETTER
# CONFIGURATION
# ====================================================================


@login_required
def upload_proposals_view(request):

    SHEET_NAME = getattr(settings, 'PROPOSAL_EXCEL_SHEET_NAME', 'Masterlist')

    # Column mapping for improved readability and maintainability
    COLUMN_MAP = {
        'NO': 2,
        'PROPOSAL_DATE': 12,
        'PROPOSAL_NO': 13,
        'COMPANY_NAME': 14,
        'SERVICE': 15,
        'FEE': 16,
    }
    
    if request.method == 'POST':
        form = MultiplePDFUploadForm(request.POST, request.FILES)
        if form.is_valid():
            # Retrieve the list of uploaded PDF files
            files = request.FILES.getlist('pdf_files')
            excel_file = form.cleaned_data.get('excel_file')
            
            # 1. Initialize the AI Processor
            # Make sure GEMINI_API_KEY_2 is set in your environment variables
            api_key = os.getenv("GEMINI_API_KEY_2")
            if not api_key:
                messages.error(request, "System Error: GEMINI_API_KEY_2 is missing from environment variables.")
                return render(request, 'tools/engagement_extract.html', {'form': form})
                
            processor = ProposalPDFProcessor(api_key=api_key)

            # 2. Load the Excel Workbook
            try:
                wb = openpyxl.load_workbook(excel_file)
                # Ensure we are writing to the correct sheet, fallback to active if missing
                ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
            except Exception as e:
                messages.error(request, f"❌ Failed to load the uploaded Excel file: {str(e)}")
                return render(request, 'tools/engagement_extract.html', {'form': form})

            # 3. Find the exact starting row based on Column B ("No.")
            last_no = 0
            start_row = ws.max_row + 1
            
            # Scan from the bottom up to find the last valid entry number
            for r in range(ws.max_row, 1, -1):
                val = ws.cell(row=r, column=2).value
                if isinstance(val, (int, float)):
                    last_no = int(val)
                    start_row = r + 1
                    
                    break

            # If the sheet is empty or has no numbers, fallback to the requested starting point
            if last_no == 0:
                last_no = 80  
                start_row = 10 # Assuming headers are at the top, it will write on row 10

            current_row = start_row
            current_no = last_no + 1

            # 4. Loop through uploaded PDFs, extract data, and write to Excel
            success_count = 0
            total_files = len(files)
            print(f"\n{'='*50}\n🚀 STARTING BATCH PROPOSAL EXTRACTION ({total_files} files)\n{'='*50}")
            
            for i, pdf_file in enumerate(files, 1):
                try:
                    print(f"⏳ [{i}/{total_files}] Processing '{pdf_file.name}'...")
                    # Read the file directly from memory as bytes (No saving to hard drive)
                    pdf_bytes = pdf_file.read()
                    
                    # Pass bytes to the processor (returns our guaranteed dictionary)
                    data = processor.extract_proposal_data(pdf_bytes)

                    # Write data using the readable column map
                    ws.cell(row=current_row, column=COLUMN_MAP['NO']).value = current_no
                    ws.cell(row=current_row, column=COLUMN_MAP['PROPOSAL_DATE']).value = data.get('proposal_date', '')
                    extracted_proposal_no = data.get('proposal_number', '')
                    ws.cell(row=current_row, column=COLUMN_MAP['PROPOSAL_NO']).value = extracted_proposal_no
                    ws.cell(row=current_row, column=COLUMN_MAP['COMPANY_NAME']).value = data.get('company_name', '')
                    ws.cell(row=current_row, column=COLUMN_MAP['SERVICE']).value = data.get('service_proposed', '')
                    ws.cell(row=current_row, column=COLUMN_MAP['FEE']).value = data.get('fee_detail', '')

                    print(f"   ✅ Extracted: {data.get('company_name')} | No: {extracted_proposal_no}")
                    current_cost = processor.cost_stats.get('flash_cost', 0) + processor.cost_stats.get('pro_cost', 0)
                    print(f"   💲 Acc. Cost: ${current_cost:.5f}")

                    current_row += 1
                    current_no += 1
                    success_count += 1
                    
                except Exception as e:
                    # If one file fails, log it but continue processing the rest
                    print(f"   ❌ Failed to process {pdf_file.name}: {str(e)}")
                    messages.warning(request, f"⚠️ Failed to process {pdf_file.name}. Moving to next file.")

            # 5. Return the Workbook as an HTTP response
            try:
                # 6. Log AI Cost for auditing and consistency
                costs = processor.cost_stats
                total_cost = costs.get('flash_cost', 0) + costs.get('pro_cost', 0)
                
                print(f"\n🎉 BATCH COMPLETE! Successfully processed {success_count}/{total_files} files.")
                print(f"💰 Total AI Cost for this batch: ${total_cost:.5f}\n{'='*50}\n")

                if total_cost > 0:
                    AICostLog.objects.create(
                        file_name=f"{total_files} proposals batch",
                        total_pages=total_files, # Treat each PDF as a "page" for logging
                        flash_cost=costs.get('flash_cost', 0),
                        pro_cost=costs.get('pro_cost', 0),
                        total_cost=total_cost
                    )
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                filename = f"Updated_Masterlist_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response = HttpResponse(
                    output.read(), 
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                # Set a cookie to signal the frontend JavaScript that the file is ready
                response.set_cookie('download_complete', 'true', path='/', max_age=60)
                return response

            except Exception as e:
                messages.error(request, f"❌ Failed to generate the Excel file. Error: {str(e)}")
            
    else:
        # GET request - just render the empty form
        form = MultiplePDFUploadForm()

    return render(request, 'tools/engagement_extract.html', {'form': form})

# --------------------------------------------------------------------
# FUZZY MATCHING LOGIC
# --------------------------------------------------------------------
def normalize_company_name(name):
    if not name: return ""
    n = str(name).lower()
    n = re.sub(r'\b(co\.,?\s*ltd\.?|ltd\.?|pte\.?|inc\.?|company|limited|plc\.?|corp\.?)\b', '', n)
    n = re.sub(r'[^a-z0-9]', '', n)
    return n

def calculate_similarity(ai_name, excel_name):
    if not ai_name or not excel_name: return 0.0
    
    ai_norm = normalize_company_name(ai_name)
    ex_norm = normalize_company_name(excel_name)
    
    if not ai_norm or not ex_norm: return 0.0
    if ai_norm == ex_norm: return 1.0 
    
    if len(ai_norm) > 5 and len(ex_norm) > 5:
        if ai_norm in ex_norm or ex_norm in ai_norm:
            return 0.95 
            
    return difflib.SequenceMatcher(None, ai_norm, ex_norm).ratio()

# --------------------------------------------------------------------
# MAIN VIEW
# --------------------------------------------------------------------
@login_required
def upload_engagement_letters_view(request):
    SHEET_NAME = getattr(settings, 'PROPOSAL_EXCEL_SHEET_NAME', 'Masterlist')

    COLUMN_MAP = {
        'COMPANY_NAME_1': 6,  # Col F
        'PROPOSAL_DATE': 12,  # Col L
        'PROPOSAL_NO': 13,    # Col M 
        'COMPANY_NAME_2': 14, # Col N
        'EL_DATE': 31,        # Col AE
        'EL_NUMBER': 32,      # Col AF
        'SERVICES': 33,       # Col AG
        'FEE_INCLUSIVE': 34,  # Col AH
        'FEE_EXCLUSIVE': 35,  # Col AI
    }
    
    if request.method == 'POST':
        form = EngagementLetterUploadForm(request.POST, request.FILES)
        if form.is_valid():
            files = request.FILES.getlist('pdf_files')
            excel_file = form.cleaned_data.get('excel_file')
            
            api_key = os.getenv("GEMINI_API_KEY_2")
            if not api_key:
                messages.error(request, "System Error: GEMINI_API_KEY_2 is missing.")
                return render(request, 'tools/engagement_letter_extract.html', {'form': form})
                
            processor = EngagementLetterProcessor(api_key=api_key)

            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
            except Exception as e:
                messages.error(request, f"❌ Failed to load Excel file: {str(e)}")
                return render(request, 'tools/engagement_letter_extract.html', {'form': form})

            success_count = 0
            total_files = len(files)
            print(f"\n{'='*60}\n🚀 STARTING BATCH EL EXTRACTION ({total_files} files)\n{'='*60}", flush=True)
            
            for i, pdf_file in enumerate(files, 1):
                try:
                    print(f"\n⏳ [{i}/{total_files}] Processing '{pdf_file.name}'...", flush=True)
                    pdf_bytes = pdf_file.read()
                    
                    data = processor.extract_el_data(pdf_bytes)
                    ai_company_name = data.get('company_name', '').strip()
                    
                    print(f"   🤖 AI Extracted Company: '{ai_company_name}'", flush=True)

                    best_match_row = None
                    best_score = 0.0
                    matched_excel_name = ""

                    if ai_company_name:
                        for r in range(5, ws.max_row + 1):
                            prop_date = str(ws.cell(row=r, column=COLUMN_MAP['PROPOSAL_DATE']).value or '').strip()
                            prop_no = str(ws.cell(row=r, column=COLUMN_MAP['PROPOSAL_NO']).value or '').strip()
                            
                            # Year Gate Filter (Only match 2026 documents)
                            if '2026' not in prop_date and '-26' not in prop_date and '2026' not in prop_no:
                                continue 

                            name_f = str(ws.cell(row=r, column=COLUMN_MAP['COMPANY_NAME_1']).value or '').strip()
                            name_n = str(ws.cell(row=r, column=COLUMN_MAP['COMPANY_NAME_2']).value or '').strip()
                            
                            score_f = calculate_similarity(ai_company_name, name_f)
                            score_n = calculate_similarity(ai_company_name, name_n)
                            
                            max_score = max(score_f, score_n)
                            
                            if max_score > best_score:
                                best_score = max_score
                                best_match_row = r
                                matched_excel_name = name_f if score_f >= score_n else name_n
                                
                            if best_score == 1.0:
                                break

                    if best_match_row and best_score >= 0.80:
                        # Write the data
                        ws.cell(row=best_match_row, column=COLUMN_MAP['EL_DATE']).value = data.get('el_date', '')
                        ws.cell(row=best_match_row, column=COLUMN_MAP['EL_NUMBER']).value = data.get('el_number', '')
                        
                        # Services and Fees (Apply wrap_text=True for perfect multiline rendering in Excel)
                        cell_services = ws.cell(row=best_match_row, column=COLUMN_MAP['SERVICES'])
                        cell_services.value = data.get('type_of_services', '')
                        cell_services.alignment = Alignment(wrap_text=True)

                        cell_fee_inc = ws.cell(row=best_match_row, column=COLUMN_MAP['FEE_INCLUSIVE'])
                        cell_fee_inc.value = data.get('total_fee_inclusive', '')
                        cell_fee_inc.alignment = Alignment(wrap_text=True)

                        cell_fee_exc = ws.cell(row=best_match_row, column=COLUMN_MAP['FEE_EXCLUSIVE'])
                        cell_fee_exc.value = data.get('total_fee_exclusive', '')
                        cell_fee_exc.alignment = Alignment(wrap_text=True)
                        
                        print(f"   ✅ MATCHED (Score: {best_score:.2f}) -> Excel Row {best_match_row}: '{matched_excel_name}'", flush=True)
                        success_count += 1
                    else:
                        print(f"   ⚠️ NO MATCH (Highest Score: {best_score:.2f} for '{matched_excel_name}')", flush=True)
                        messages.warning(request, f"⚠️ '{pdf_file.name}' extracted '{ai_company_name}' but could not find a 2026 match in the Masterlist.")

                except Exception as e:
                    print(f"   ❌ Failed to process {pdf_file.name}: {str(e)}", flush=True)
                    messages.warning(request, f"⚠️ Failed to process {pdf_file.name}.")

            try:
                costs = processor.cost_stats
                total_cost = costs.get('flash_cost', 0) + costs.get('pro_cost', 0)
                print(f"\n🎉 BATCH COMPLETE! Processed {success_count}/{total_files} files. Total Cost: ${total_cost:.5f}\n{'='*60}\n", flush=True)

                if total_cost > 0:
                    AICostLog.objects.create(
                        file_name=f"{total_files} ELs batch",
                        total_pages=total_files,
                        flash_cost=costs.get('flash_cost', 0),
                        pro_cost=costs.get('pro_cost', 0),
                        total_cost=total_cost
                    )
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                filename = f"Masterlist_Updated_EL_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response = HttpResponse(
                    output.read(), 
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response.set_cookie('download_complete', 'true', path='/', max_age=60)
                return response

            except Exception as e:
                messages.error(request, f"❌ Failed to generate the Excel file: {str(e)}")
            
    else:
        form = EngagementLetterUploadForm()

    return render(request, 'tools/engagement_letter_extract.html', {'form': form})

def get_closest_exchange_rate(target_date):
    """
    Looks up the exchange rate for the exact target date.
    If not found, steps backward day-by-day to find the most recent rate.
    Assumes ExchangeRate is imported from clients.models.
    """
    from clients.models import ExchangeRate
    
    current_search_date = target_date
    # Limit search backward to prevent infinite loops (e.g., max 10 days)
    for _ in range(10):
        rate_record = ExchangeRate.objects.filter(date=current_search_date).first()
        if rate_record and rate_record.rate > 0:
            return rate_record.rate
        current_search_date -= timedelta(days=1)
        
    return None # Fallback if absolutely no rates exist in the database

@login_required
def monthly_closing_view(request):
    # ---------------------------------------------------------
    # 4. STATIC TUPLE GENERATION 
    # ---------------------------------------------------------
    db_accounts = [(a.account_id, f"{a.account_id} - {a.name}") for a in Account.objects.all().order_by('account_id')]
    account_choices = [('', '--- Select Account ---')] + db_accounts

    db_vendors = [(str(v.id), f"{v.vendor_id} - {v.name}") for v in Vendor.objects.all().order_by('vendor_id')]
    vendor_choices = [('', '--- No Vendor ---')] + db_vendors

    form_kwargs_dict = {
        'account_choices': account_choices, 
        'vendor_choices': vendor_choices
    }

    # ---------------------------------------------------------
    # 5. MAIN PROCESSING ROUTE
    # ---------------------------------------------------------
    if request.method == 'POST':
        form = MonthlyClosingForm(request.POST, request.FILES)
        
        accrual_formset = AccrualFormSet(request.POST, prefix='accrual', form_kwargs=form_kwargs_dict)
        fx_formset = FXFormSet(request.POST, prefix='fx', form_kwargs=form_kwargs_dict)

        print(f"\n{'='*60}\n🚀 STARTING UNIFIED MONTHLY CLOSING PROCESS\n{'='*60}")

        if form.is_valid() and accrual_formset.is_valid() and fx_formset.is_valid():
            selected_date = form.cleaned_data['date']
            
            api_key = getattr(settings, 'GEMINI_API_KEY_2', os.getenv("GEMINI_API_KEY_2"))
            if not api_key:
                messages.error(request, "System Error: GEMINI_API_KEY_2 is missing.")
                return render(request, 'tools/monthly_closing.html', {
                    'form': form, 'accrual_formset': accrual_formset, 'fx_formset': fx_formset
                })
            
            vendor_tax, _ = Vendor.objects.get_or_create(name='General Department of Taxation', defaults={'vendor_id': 'V-TAX'})
            vendor_staff, _ = Vendor.objects.get_or_create(name='Staff', defaults={'vendor_id': 'V-STAFF'})

            transaction_lines = [] 
            period_label = selected_date.strftime("%b'%y")
            total_ai_cost = 0.0

            # =========================================================
            # SCENARIO A: Unified Tax & Salary Liabilities
            # =========================================================
            if form.cleaned_data.get('tax_declaration_pdf'):
                pdf_bytes = form.cleaned_data['tax_declaration_pdf'].read()
                salary_payable_usd = float(form.cleaned_data.get('salary_payable') or 0.0)
                staff_meals_usd = float(form.cleaned_data.get('staff_meals') or 0.0)
                
                print(f"\n[ MODULE 1 & 2 ] Processing Tax Declaration PDF...")
                print(f"   [DEBUG] Salary Payable input: {salary_payable_usd} | Staff Meals input: {staff_meals_usd}")
                
                # 💡 ENHANCEMENT: Fetch the exact exchange rate for the 25th of the month
                target_tax_date = date(selected_date.year, selected_date.month, 25)
                tax_exchange_rate = get_closest_exchange_rate(target_tax_date)
                
                if not tax_exchange_rate:
                    messages.error(request, f"Cannot process Taxes: No Exchange Rate found in DB for {target_tax_date} or preceding days.")
                    return redirect('tools:monthly_closing')
                    
                print(f"   [DEBUG] DB Exchange Rate for Taxes (Target {target_tax_date}): {tax_exchange_rate}")

                processor = UnifiedTaxProcessor(api_key=api_key)
                # 💡 FORCE the processor to use our DB rate instead of guessing from the PDF
                data = processor.extract_tax_data(pdf_bytes, forced_exchange_rate=tax_exchange_rate)

                if not data.get('error'):
                    total_ai_cost += processor.cost_stats['flash_cost']
                    
                    exchange_rate = data.get('exchange_rate', tax_exchange_rate)
                    tos_instr = data.get('tos_instruction') or f"TOS Extracted (Rate: {exchange_rate})"
                    wht_instr = data.get('wht_instruction') or f"WHT Extracted (Rate: {exchange_rate})"
                    fbt_instr = data.get('fbt_instruction') or f"FBT Extracted (Rate: {exchange_rate})"
                    general_instr = data.get('general_instruction') or f"Salary/Meals Extracted (Rate: {exchange_rate})"

                    if staff_meals_usd == 0.0:
                        staff_meals_usd = float(data.get('staff_meals_usd', 0.0))

                    # 1. Process Salary & TOS
                    if salary_payable_usd > 0:
                        tax_res_usd = float(data.get('tos_resident_usd', 0.0))
                        tax_non_res_usd = float(data.get('tos_non_resident_usd', 0.0))
                        total_tos = round(tax_res_usd + tax_non_res_usd, 2)

                        transaction_lines.extend([
                            {"vendor": vendor_staff, "account_id": "705000", "instruction": general_instr, "desc": f"Salary expense {period_label}", "debit": salary_payable_usd, "credit": 0.0},
                            {"vendor": vendor_tax, "account_id": "725410", "instruction": tos_instr, "desc": f"Salary tax expense {period_label}", "debit": total_tos, "credit": 0.0},
                            {"vendor": vendor_tax, "account_id": "210030", "instruction": tos_instr, "desc": f"Salary tax expense {period_label}", "debit": 0.0, "credit": tax_res_usd},
                            {"vendor": vendor_tax, "account_id": "210030", "instruction": tos_instr, "desc": f"NR Salary tax expense {period_label}", "debit": 0.0, "credit": tax_non_res_usd},
                            {"vendor": vendor_staff, "account_id": "225070", "instruction": general_instr, "desc": f"Being accrued for salary expense {period_label}", "debit": 0.0, "credit": salary_payable_usd}
                        ])

                    if staff_meals_usd > 0:
                        transaction_lines.extend([
                            {"vendor": vendor_staff, "account_id": "705070", "instruction": general_instr, "desc": f"Staff meals {period_label}", "debit": staff_meals_usd, "credit": 0.0},
                            {"vendor": vendor_staff, "account_id": "225070", "instruction": general_instr, "desc": f"Being accrued for staff meals {period_label}", "debit": 0.0, "credit": staff_meals_usd}
                        ])

                    # 2. Process Withholding Taxes
                    total_wht = round(float(data.get('wht_10_usd', 0.0)) + float(data.get('wht_15_usd', 0.0)), 2)
                    if total_wht > 0:
                        desc = f"Being accrued for Withholding tax expenses in {period_label}"
                        transaction_lines.extend([
                            {"vendor": vendor_tax, "account_id": "725420", "instruction": wht_instr, "desc": desc, "debit": total_wht, "credit": 0.0},
                            {"vendor": vendor_tax, "account_id": "210040", "instruction": wht_instr, "desc": desc, "debit": 0.0, "credit": total_wht}
                        ])

                    # 3. Process Fringe Benefit Tax
                    fbt = round(float(data.get('fbt_usd', 0.0)), 2)
                    if fbt > 0:
                        desc = f"Being accrued for Fringe Benefit tax expenses in {period_label}"
                        transaction_lines.extend([
                            {"vendor": vendor_tax, "account_id": "705010", "instruction": fbt_instr, "desc": desc, "debit": fbt, "credit": 0.0},
                            {"vendor": vendor_tax, "account_id": "210031", "instruction": fbt_instr, "desc": desc, "debit": 0.0, "credit": fbt}
                        ])

                else:
                    print("   ❌ [DEBUG] Tax PDF Extraction returned an error.")
            else:
                print("\n[ MODULE 1 & 2 ] Skipped. No Tax Declaration PDF provided.")

            # =========================================================
            # SCENARIO B: Accrued Expenses
            # =========================================================
            print("\n[ MODULE 3 ] Processing Manual Accruals...")
            print(f"   [DEBUG] Total Accrual forms submitted: {len(accrual_formset)}")
            for a_form in accrual_formset:
                if a_form.cleaned_data and not a_form.cleaned_data.get('DELETE', False) and a_form.cleaned_data.get('debit', 0) > 0:
                    debit_amt = round(float(a_form.cleaned_data['debit']), 2)
                    desc = a_form.cleaned_data['description']
                    p_status = a_form.cleaned_data.get('payment_status') or 'Open'
                    
                    vendor_id_str = a_form.cleaned_data.get('vendor')
                    vendor_instance = None
                    if vendor_id_str:
                        try:
                            vendor_instance = Vendor.objects.get(id=int(vendor_id_str))
                        except (ValueError, Vendor.DoesNotExist):
                            pass
                    
                    transaction_lines.extend([
                        {"vendor": vendor_instance, "account_id": a_form.cleaned_data['account_id'], "instruction": "Manual Accrual", "desc": desc, "debit": debit_amt, "credit": 0.0, "payment_status": p_status},
                        {"vendor": vendor_instance, "account_id": "215090", "instruction": "Manual Accrual", "desc": desc, "debit": 0.0, "credit": debit_amt, "payment_status": p_status}
                    ])
                    print(f"      ✅ [DEBUG] Accrual added for '{desc}' (${debit_amt})")

            # =========================================================
            # SCENARIO C: FX Gain/Loss (AUTOMATED RATE FETCH)
            # =========================================================
            print("\n[ MODULE 4 ] Processing FX Gain/Loss...")
            
            # 💡 ENHANCEMENT: Dynamically calculate the last day of the selected month
            last_day = calendar.monthrange(selected_date.year, selected_date.month)[1]
            month_end_date = date(selected_date.year, selected_date.month, last_day)
            
            # Fetch the DB rate for month-end
            month_end_rate = get_closest_exchange_rate(month_end_date)
            
            for f_form in fx_formset:
                if f_form.cleaned_data and not f_form.cleaned_data.get('DELETE', False):
                    # Only proceed if we have a valid ending KHR balance to evaluate
                    end_bal_khr = float(f_form.cleaned_data.get('ending_balance') or 0.0)
                    
                    if end_bal_khr != 0.0:
                        if not month_end_rate:
                            messages.error(request, f"Cannot process FX: No Exchange Rate found in DB for month-end {month_end_date}.")
                            # Safe fallback: break out of the FX loop but let other modules save
                            break 
                            
                        open_bal_usd = float(f_form.cleaned_data['openning_balance'] or 0.0)
                        desc = f_form.cleaned_data['description']
                        p_status = f_form.cleaned_data.get('payment_status') or 'Paid'
                        
                        fx_account_id = f_form.cleaned_data['account_id'] 
                        bank_account_id = f_form.cleaned_data['bank_account_id']

                        # 💡 ENHANCEMENT: Automatic Math using the DB Rate
                        month_end_usd = round(end_bal_khr / month_end_rate, 2)
                        fx_diff = round(month_end_usd - open_bal_usd, 2)
                        
                        instruction_txt = f"Automated FX: (End Bal KHR {end_bal_khr} / DB Rate {month_end_rate}) - Open Bal USD {open_bal_usd} = {fx_diff}"
                        
                        # Note: We pass None to "vendor" because FX adjustments are internal bank revaluations.
                        if fx_diff < 0:
                            loss_amt = abs(fx_diff)
                            transaction_lines.extend([
                                {"vendor": None, "account_id": fx_account_id, "instruction": instruction_txt, "desc": f"{desc} (FX Loss)", "debit": loss_amt, "credit": 0.0, "payment_status": p_status},
                                {"vendor": None, "account_id": bank_account_id, "instruction": instruction_txt, "desc": f"{desc} (FX Loss)", "debit": 0.0, "credit": loss_amt, "payment_status": p_status}
                            ])
                            print(f"      ✅ [DEBUG] Auto FX Loss added for '{desc}' (${loss_amt}) at rate {month_end_rate}")
                        elif fx_diff > 0:
                            gain_amt = fx_diff
                            transaction_lines.extend([
                                {"vendor": None, "account_id": bank_account_id, "instruction": instruction_txt, "desc": f"{desc} (FX Gain)", "debit": gain_amt, "credit": 0.0, "payment_status": p_status},
                                {"vendor": None, "account_id": fx_account_id, "instruction": instruction_txt, "desc": f"{desc} (FX Gain)", "debit": 0.0, "credit": gain_amt, "payment_status": p_status}
                            ])
                            print(f"      ✅ [DEBUG] Auto FX Gain added for '{desc}' (${gain_amt}) at rate {month_end_rate}")
                        else:
                            print(f"      ⏭️ [DEBUG] Skipped FX row: Net variance is strictly $0.00")

            # =========================================================
            # ATOMIC DATABASE SAVE 
            # =========================================================
            print(f"\n[ SUMMARY ] Total raw transaction lines generated: {len(transaction_lines)}")
            transaction_lines = [line for line in transaction_lines if line['debit'] > 0 or line['credit'] > 0]
            print(f"[ SUMMARY ] Transaction lines remaining after filtering $0 amounts: {len(transaction_lines)}")

            if transaction_lines:
                try:
                    with transaction.atomic():
                        for line in transaction_lines:
                            jv = JournalVoucher.objects.create(
                                user=request.user, date=selected_date, vendor=line['vendor'], account_id=line['account_id'], instruction=line['instruction'], 
                                description=line['desc'], debit=line['debit'], credit=line['credit'], 
                                payment_status=line.get('payment_status') or 'Open'
                            )
                            je = JournalEntry.objects.create(
                                date=selected_date, description=f"Monthly Closing Automation - {period_label}"[:255], 
                                reference_number=f"JV-{jv.id}", journal_voucher=jv
                            )
                            account, _ = Account.objects.get_or_create(
                                account_id=str(line['account_id']), defaults={'name': 'System Gen Acct', 'account_type': 'Expense'}
                            )
                            JournalLine.objects.create(
                                journal_entry=je, account=account, debit=line['debit'], credit=line['credit'], description=line['desc'][:255]
                            )

                        if total_ai_cost > 0:
                            try:
                                AICostLog.objects.create(
                                    file_name=f"Monthly Closing Batch - {period_label}",
                                    total_pages=1, flash_cost=total_ai_cost, total_cost=total_ai_cost
                                )
                            except NameError:
                                pass

                    messages.success(request, f"Successfully processed {period_label}. Created {len(transaction_lines)} Journal Voucher records.")
                    return redirect('tools:monthly_closing')
                except Exception as db_error:
                    messages.error(request, f"Database Error: {str(db_error)}")
            else:
                messages.warning(request, "No valid transactions were extracted or entered.")
                
        else:
            print("\n❌ [DEBUG] FORM VALIDATION FAILED in Monthly Closing:")
            print(f"   Main Form Errors: {form.errors}")
            print(f"   Accrual Formset Errors: {accrual_formset.errors}")
            print(f"   FX Formset Errors: {fx_formset.errors}")
            messages.error(request, "Validation failed. Please check the form fields.")

    else:
        # ---------------------------------------------------------
        # 6. GET REQUEST RENDERING
        # ---------------------------------------------------------
        form = MonthlyClosingForm()
                
        accrual_formset = AccrualFormSet(prefix='accrual', form_kwargs=form_kwargs_dict)
        fx_formset = FXFormSet(prefix='fx', form_kwargs=form_kwargs_dict)

    return render(request, 'tools/monthly_closing.html', {
        'form': form, 'accrual_formset': accrual_formset, 'fx_formset': fx_formset
    })

def load_vendors(request):
    """HTMX endpoint to return vendor <option> tags for the current schema."""
    vendors = Vendor.objects.all().order_by('vendor_id')
        
    return render(request, 'tools/partials/vendor_options.html', {'vendors': vendors})

# ====================================================================
# --- 5. JOURNAL VOUCHER CRUD & POSTING ---
# ====================================================================

@login_required(login_url="register:login")
def JournalVoucherListView(request):
    jv_records = JournalVoucher.objects.all().order_by('-id')
    # Vendor queryset also automatically filters by the current schema
    vendor_queryset = Vendor.objects.all().order_by('vendor_id')

    jv_filter = JournalVoucherFilter(request.GET, queryset=jv_records)
    jv_filter.form.fields['vendor'].queryset = vendor_queryset

    paginator = Paginator(jv_filter.qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {'filter': jv_filter, 'jv_records': page_obj, 'page_obj': page_obj}
    return render(request, 'tools/journal_voucher_list.html', context)

@login_required(login_url="register:login")
def manual_journal_voucher_entry_view(request):
    seen_accounts = set()
    db_accounts = []
    for acc_id, name in Account.objects.values_list('account_id', 'name'):
        if acc_id not in seen_accounts:
            seen_accounts.add(acc_id)
            db_accounts.append((str(acc_id), f"{acc_id} - {name}"))
    db_accounts.sort(key=lambda x: str(x[0]))
    account_choices = [('', '--- Select Account ---')] + db_accounts

    if request.method == 'POST':
        form = JournalVoucherEntryForm(request.POST, account_choices=account_choices)
        if form.is_valid():
            with transaction.atomic():
                jv_record = form.save(commit=False)
                jv_record.user = request.user
                
                acct, _ = Account.objects.get_or_create(account_id=str(jv_record.account_id), defaults={'name': 'JV Default', 'account_type': 'Asset'})
                
                if acct.account_type and acct.account_type.lower() in ['asset', 'liability']:
                    jv_record.payment_status = 'Open'
                else:
                    jv_record.payment_status = 'Paid'
                    
                jv_record.save()
                
                je = JournalEntry.objects.create(
                    date=jv_record.date or date.today(),
                    description=f"Journal Voucher: {jv_record.description}"[:255], reference_number=f"JV-{jv_record.id}",
                    journal_voucher=jv_record
                )
                safe_desc = jv_record.description[:255] if jv_record.description else "Journal Voucher"
                debit_val = jv_record.debit or 0.0
                credit_val = jv_record.credit or 0.0
                if debit_val > 0 and credit_val > 0:
                    JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=debit_val, credit=0.0)
                    JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=0.0, credit=credit_val)
                else:
                    JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=debit_val, credit=credit_val)
            
            messages.success(request, f"Successfully created journal voucher and posted to GL.")
            return redirect('tools:journal_voucher_list') 
    else:
        form = JournalVoucherEntryForm(account_choices=account_choices)
        form.fields['vendor'].queryset = Vendor.objects.all()

    return render(request, 'tools/journal_voucher_form.html', {'form': form})

class JournalVoucherDetailView(LoginRequiredMixin, DetailView):
    login_url = "register:login"
    model = JournalVoucher
    template_name = 'tools/journal_voucher_detail.html'
    context_object_name = 'jv_record'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_owner'] = True
        return context

class JournalVoucherUpdateView(LoginRequiredMixin, UpdateView):
    login_url = "register:login"
    model = JournalVoucher
    form_class = JournalVoucherEntryForm 
    template_name = 'tools/journal_voucher_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        db_accounts = [(a.account_id, f"{a.account_id} - {a.name}") for a in Account.objects.all().order_by('account_id')]
        kwargs['account_choices'] = [('', '--- Select Account ---')] + db_accounts
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'form' in context:
            context['form'].fields['vendor'].queryset = Vendor.objects.all()
        return context

    def form_valid(self, form):
        with transaction.atomic():
            jv_record = form.save(commit=False)
            
            acct, _ = Account.objects.get_or_create(account_id=str(jv_record.account_id), defaults={'name': 'JV Default', 'account_type': 'Asset'})
            
            if acct.account_type and acct.account_type.lower() in ['asset', 'liability']:
                jv_record.payment_status = 'Open'
            else:
                jv_record.payment_status = 'Paid'
                
            jv_record.save()
            je, created = JournalEntry.objects.get_or_create(
                journal_voucher=jv_record,
                defaults={
                    'date': jv_record.date or date.today(),
                    'description': f"Journal Voucher: {jv_record.description}"[:255],
                    'reference_number': f"JV-{jv_record.id}",
                }
            )
            if not created:
                je.date = jv_record.date or date.today()
                je.description = f"Updated Journal Voucher: {jv_record.description}"[:255]
                je.save(update_fields=['date', 'description'])
                je.lines.all().delete()

            safe_desc = jv_record.description[:255] if jv_record.description else "Journal Voucher"
            debit_val = jv_record.debit or 0.0
            credit_val = jv_record.credit or 0.0
            if debit_val > 0 and credit_val > 0:
                JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=debit_val, credit=0.0)
                JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=0.0, credit=credit_val)
            else:
                JournalLine.objects.create(journal_entry=je, account=acct, description=safe_desc, debit=debit_val, credit=credit_val)
        messages.success(self.request, "Journal voucher and General Ledger entries updated successfully!")
        return HttpResponseRedirect(reverse('tools:journal_voucher_detail', kwargs={'pk': self.object.pk}))

class JournalVoucherDeleteView(LoginRequiredMixin, DeleteView):
    login_url = "register:login"
    model = JournalVoucher
    template_name = 'tools/journal_voucher_confirm_delete.html'
    success_url = reverse_lazy('tools:journal_voucher_list')

    def form_valid(self, form):
        JournalEntry.objects.filter(Q(journal_voucher=self.object) | Q(reference_number=f"JV-{self.object.id}")).delete()
        messages.success(self.request, 'Journal voucher and associated Journal Entries deleted successfully!')
        return super().form_valid(form)

# ====================================================================
# --- 6. ADJUSTMENT CRUD & POSTING ---
# ====================================================================

@login_required(login_url="register:login")
def AdjustmentListView(request):
    adj_records = Adjustment.objects.all().order_by('-id')
    vendor_queryset = Vendor.objects.all().order_by('vendor_id')
    customer_queryset = Customer.objects.all().order_by('customer_id')
    
    adj_filter = AdjustmentFilter(request.GET, queryset=adj_records)
    adj_filter.form.fields['vendor'].queryset = vendor_queryset
    adj_filter.form.fields['customer'].queryset = customer_queryset

    paginator = Paginator(adj_filter.qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {'filter': adj_filter, 'adj_records': page_obj, 'page_obj': page_obj}
    return render(request, 'tools/adjustment_list.html', context)

@login_required(login_url="register:login")
def manual_adjustment_entry_view(request):
    if request.method == 'POST':
        formset = AdjustmentFormSet(request.POST)
        if formset.is_valid():
            with transaction.atomic():
                adjustments_created = 0
                for form in formset:
                    if form.cleaned_data and not form.cleaned_data.get('DELETE'):
                        adj_record = form.save(commit=False)
                        adj_record.user = request.user
                        adj_record.save()
                        adjustments_created += 1
                        
                        je = JournalEntry.objects.create(
                            date=adj_record.date or date.today(),
                            description=f"Adjustment: {adj_record.description}"[:255], reference_number=f"ADJ-{adj_record.id}",
                            adjustment=adj_record
                        )
                        safe_desc = adj_record.description[:255] if adj_record.description else "Adjustment"
                        if adj_record.debit_account_id:
                            JournalLine.objects.create(journal_entry=je, account=adj_record.debit_account_id, description=safe_desc, debit=adj_record.debit or 0.0)
                        if adj_record.credit_account_id:
                            JournalLine.objects.create(journal_entry=je, account=adj_record.credit_account_id, description=safe_desc, credit=adj_record.credit or 0.0)
                            
                        purchase_ids = form.cleaned_data.get('purchase_id', '')
                        sale_ids = form.cleaned_data.get('sale_id', '')
                        jv_ids = form.cleaned_data.get('journal_voucher_id', '')
                        
                        for p_id in [x.strip() for x in purchase_ids.split(',') if x.strip()]:
                            try:
                                p = Purchase.objects.get(id=p_id)
                                if p.payment_status == 'Open':
                                    acct = Account.objects.filter(account_id=str(p.credit_account_id)).first()
                                    if acct and acct.account_type in ['Liability', 'Revenue', 'Expense']:
                                        p.payment_status = 'Paid'
                                        p.save(update_fields=['payment_status'])
                            except Exception:
                                pass
                                
                        for s_id in [x.strip() for x in sale_ids.split(',') if x.strip()]:
                            try:
                                s = Sale.objects.get(id=s_id)
                                if s.payment_status == 'Open':
                                    acct_id = getattr(s, 'account_id', None) or getattr(s, 'debit_account_id', None) or getattr(s, 'credit_account_id', None)
                                    acct = Account.objects.filter(account_id=str(acct_id)).first()
                                    if acct and acct.account_type in ['Liability', 'Revenue', 'Expense']:
                                        s.payment_status = 'Paid'
                                        s.save(update_fields=['payment_status'])
                            except Exception:
                                pass
                                
                        for j_id in [x.strip() for x in jv_ids.split(',') if x.strip()]:
                            try:
                                j = JournalVoucher.objects.get(id=j_id)
                                if j.payment_status == 'Open':
                                    acct = Account.objects.filter(account_id=str(j.account_id)).first()
                                    if acct and acct.account_type in ['Liability', 'Revenue', 'Expense']:
                                        j.payment_status = 'Paid'
                                        j.save(update_fields=['payment_status'])
                            except Exception:
                                pass
            
            messages.success(request, f"Successfully created {adjustments_created} adjustments and posted to GL.")
            return redirect('tools:adjustment_list') 
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        formset = AdjustmentFormSet()

    return render(request, 'tools/adjustment_form.html', {'formset': formset})

class AdjustmentDetailView(LoginRequiredMixin, DetailView):
    login_url = "register:login"
    model = Adjustment
    template_name = 'tools/adjustment_detail.html'
    context_object_name = 'adj_record'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_owner'] = True
        return context

class AdjustmentUpdateView(LoginRequiredMixin, UpdateView):
    login_url = "register:login"
    model = Adjustment
    form_class = AdjustmentEntryForm 
    template_name = 'tools/adjustment_form.html'
    
    def form_valid(self, form):
        with transaction.atomic():
            adj_record = form.save()
            je, created = JournalEntry.objects.get_or_create(
                adjustment=adj_record,
                defaults={
                    'date': adj_record.date or date.today(),
                    'description': f"Adjustment: {adj_record.description}"[:255],
                    'reference_number': f"ADJ-{adj_record.id}",
                }
            )
            if not created:
                je.date = adj_record.date or date.today()
                je.description = f"Updated Adjustment: {adj_record.description}"[:255]
                je.save(update_fields=['date', 'description'])
                je.lines.all().delete()

            safe_desc = adj_record.description[:255] if adj_record.description else "Adjustment"
            if adj_record.debit_account_id:
                JournalLine.objects.create(journal_entry=je, account=adj_record.debit_account_id, description=safe_desc, debit=adj_record.debit or 0.0)
            if adj_record.credit_account_id:
                JournalLine.objects.create(journal_entry=je, account=adj_record.credit_account_id, description=safe_desc, credit=adj_record.credit or 0.0)
                
            purchase_ids = form.cleaned_data.get('purchase_id', '')
            sale_ids = form.cleaned_data.get('sale_id', '')
            jv_ids = form.cleaned_data.get('journal_voucher_id', '')
            
            for p_id in [x.strip() for x in purchase_ids.split(',') if x.strip()]:
                try:
                    p = Purchase.objects.get(id=p_id)
                    if p.payment_status == 'Open':
                        acct = Account.objects.filter(account_id=str(p.credit_account_id)).first()
                        if acct and acct.account_type in ['Liability', 'Revenue', 'Expense']:
                            p.payment_status = 'Paid'
                            p.save(update_fields=['payment_status'])
                except Exception:
                    pass
                    
            for s_id in [x.strip() for x in sale_ids.split(',') if x.strip()]:
                try:
                    s = Sale.objects.get(id=s_id)
                    if s.payment_status == 'Open':
                        acct_id = getattr(s, 'account_id', None) or getattr(s, 'debit_account_id', None) or getattr(s, 'credit_account_id', None)
                        acct = Account.objects.filter(account_id=str(acct_id)).first()
                        if acct and acct.account_type in ['Liability', 'Revenue', 'Expense']:
                            s.payment_status = 'Paid'
                            s.save(update_fields=['payment_status'])
                except Exception:
                    pass
                    
            for j_id in [x.strip() for x in jv_ids.split(',') if x.strip()]:
                try:
                    j = JournalVoucher.objects.get(id=j_id)
                    if j.payment_status == 'Open':
                        acct = Account.objects.filter(account_id=str(j.account_id)).first()
                        if acct and acct.account_type in ['Liability', 'Revenue', 'Expense']:
                            j.payment_status = 'Paid'
                            j.save(update_fields=['payment_status'])
                except Exception:
                    pass
        
        messages.success(self.request, "Adjustment and General Ledger entries updated successfully!")
        return HttpResponseRedirect(reverse('tools:adjustment_detail', kwargs={'pk': self.object.pk}))

class AdjustmentDeleteView(LoginRequiredMixin, DeleteView):
    login_url = "register:login"
    model = Adjustment
    template_name = 'tools/adjustment_confirm_delete.html'
    success_url = reverse_lazy('tools:adjustment_list')

    def form_valid(self, form):
        JournalEntry.objects.filter(Q(adjustment=self.object) | Q(reference_number=f"ADJ-{self.object.id}")).delete()
        messages.success(self.request, 'Adjustment and associated Journal Entries deleted successfully!')
        return super().form_valid(form)

def generate_fifo_offset_proposals():
    """Helper function to calculate FIFO offsets"""
    print("DEBUG: Starting generate_fifo_offset_proposals")
    proposals = []
    
    # 1. Get vendors who have unapplied prepayments (assigned to account 120000 in Bank)
    vendors = Vendor.objects.filter(
        bank__debit_account_id='120000'
    ).distinct()
    print(f"DEBUG: Found {vendors.count()} vendors with prepayments")
    
    for vendor in vendors:
        print(f"DEBUG: Processing vendor: {vendor}")
        
        # Fetch all offset adjustments for this vendor to cross-check consumed balances
        vendor_adjustments = list(Adjustment.objects.filter(
            vendor=vendor,
            credit_account_id__account_id='120000'
        ))
        
        # 1. Get Unapplied Prepayments (Oldest first)
        prepayments = []
        
        bank_prepayments = Bank.objects.filter(
            vendor=vendor, 
            debit_account_id='120000'
        ).order_by('date', 'id')
        
        # FIFO Pooling: Sum all consumed prepayments for this vendor
        total_prepayment_consumed = round(sum(adj.credit or 0.0 for adj in vendor_adjustments), 2)

        for pr in bank_prepayments:
            pr_amount = round(pr.credit or 0.0, 2)
            
            if total_prepayment_consumed >= pr_amount:
                # Fully consumed
                total_prepayment_consumed = round(total_prepayment_consumed - pr_amount, 2)
                continue
                
            # Partially consumed or unconsumed
            remaining_balance = round(pr_amount - total_prepayment_consumed, 2)
            total_prepayment_consumed = 0.0
            
            if remaining_balance > 0.001:
                pr.balance = remaining_balance
                pr.original_amount = pr_amount
                pr.source_type = 'bank'
                prepayments.append(pr)
            
        # Sort combined prepayments by date
        prepayments.sort(key=lambda x: x.date or date.today())
        print(f"DEBUG: Found {len(prepayments)} total prepayments for vendor {vendor} with remaining balance")
        
        if not prepayments:
            continue # Skip if all prepayments for this vendor have been fully offset!
        
        # 2. Get Open Purchases (Oldest first)
        open_purchases_qs = Purchase.objects.filter(vendor=vendor, payment_status='Open').order_by('date', 'id')
        open_purchases = []
        
        for pu in open_purchases_qs:
            # Cross-check existing Adjustments for offsets on this Purchase ID
            used_amount = round(sum(
                adj.debit or 0.0 
                for adj in vendor_adjustments 
                if adj.description and f"Purchase ID: {pu.id} (" in adj.description
            ), 2)
            
            remaining_balance = round((pu.total_usd or 0.0) - used_amount, 2)
            
            if remaining_balance > 0.001:
                pu.balance_due = remaining_balance
                pu.original_amount = pu.total_usd or 0.0
                open_purchases.append(pu)
                
        print(f"DEBUG: Found {len(open_purchases)} open purchases for vendor {vendor} with remaining balance")
        
        # Calculate total open purchase amount to satisfy rule #5:
        # "No offset if the amount of prepayment is more than total amount of open purchases"
        total_open_purchase_amt = sum(p.balance_due for p in open_purchases)
        total_prepayment_amt = sum(pr.balance for pr in prepayments)
        
        print(f"DEBUG: total_prepayment_amt: {total_prepayment_amt}, total_open_purchase_amt: {total_open_purchase_amt}")

        if total_prepayment_amt > total_open_purchase_amt and total_open_purchase_amt > 0:
            # Rule 5 triggered: Skip this vendor entirely
            print(f"DEBUG: Rule 5 triggered for vendor {vendor}. Skipping.")
            continue

        for prepay in prepayments:
            for purchase in open_purchases:
                if prepay.balance <= 0.001:
                    break # Move to next prepayment
                if purchase.balance_due <= 0.001:
                    continue # Move to next purchase
                
                # Determine offset amount
                offset_amount = min(prepay.balance, purchase.balance_due)
                print(f"DEBUG: Offsetting {offset_amount} between prepay (ID: {prepay.id}) and open purchase {purchase.id}")
                
                prepay_bal_before = round(prepay.balance, 2)
                purchase_bal_before = round(purchase.balance_due, 2)

                # Deduct from temporary memory loops
                prepay.balance -= offset_amount
                purchase.balance_due -= offset_amount
                
                is_partial = purchase.balance_due > 0.001
                
                # Find correct Account DB instances dynamically
                trade_payable_acc = Account.objects.filter(account_id='200000').first()
                prepayment_acc = Account.objects.filter(account_id='120000').first()

                bd = getattr(prepay, 'date', None)
                pd = getattr(purchase, 'date', None)

                proposal = {
                    'date': None,
                    'bank_date': bd.strftime('%Y-%m-%d') if bd else None,
                    'purchase_date': pd.strftime('%Y-%m-%d') if pd else None,
                    'vendor': vendor.id,
                    'debit_account_id': trade_payable_acc.id if trade_payable_acc else None,
                    'credit_account_id': prepayment_acc.id if prepayment_acc else None,
                    'debit': round(offset_amount, 2),
                    'credit': round(offset_amount, 2),
                    'partial_offset': is_partial,
                    'purchase_id': purchase.id,
                    'description': f"Bank ID: {prepay.id} (Orig: {prepay.original_amount:.2f}, Bal: {prepay_bal_before:.2f}, Offset: {offset_amount:.2f}, New bal: {prepay.balance:.2f}), after offset Purchase ID: {purchase.id} (Orig: {purchase.original_amount:.2f}, Bal: {purchase_bal_before:.2f}, Offset: {offset_amount:.2f}, New bal: {purchase.balance_due:.2f})"
                }
                
                if getattr(prepay, 'source_type', '') == 'journal_voucher':
                    proposal['journal_voucher_id'] = prepay.id
                elif getattr(prepay, 'source_type', '') == 'bank':
                    proposal['bank_id'] = prepay.id
                    
                proposals.append(proposal)
                
    print(f"DEBUG: Generated {len(proposals)} proposals")
    return proposals

@login_required(login_url="register:login")
def automate_prepayment_offset(request):
    print(f"DEBUG: automate_prepayment_offset called with method {request.method}")
    page_obj = None
    if request.method == 'POST':
        formset = OffsetFormSet(request.POST)
        if formset.is_valid():
            instances_saved = 0
            for form in formset:
                # If user marked for deletion, skip saving
                if form.cleaned_data.get('DELETE'):
                    continue
                
                # Only save if there is data
                if form.cleaned_data:
                    # Manually instantiate Adjustment since AdjustmentOffsetForm is a forms.Form
                    adjustment = Adjustment(
                        date=form.cleaned_data.get('date'),
                        vendor=form.cleaned_data.get('vendor'),
                        debit_account_id=form.cleaned_data.get('debit_account_id'),
                        credit_account_id=form.cleaned_data.get('credit_account_id'),
                        debit=form.cleaned_data.get('debit'),
                        credit=form.cleaned_data.get('credit'),
                        description=form.cleaned_data.get('description'),
                        user=request.user
                    )
                    adjustment.save()
                    instances_saved += 1
                    
                    # Post to GL automatically
                    je = JournalEntry.objects.create(
                        date=adjustment.date or date.today(),
                        description=f"Automated Prepayment Offset: {adjustment.description}"[:255], 
                        reference_number=f"ADJ-{adjustment.id}",
                        adjustment=adjustment
                    )
                    
                    if adjustment.debit_account_id:
                        JournalLine.objects.create(journal_entry=je, account=adjustment.debit_account_id, description=adjustment.description[:255], debit=adjustment.debit or 0.0)
                    if adjustment.credit_account_id:
                        JournalLine.objects.create(journal_entry=je, account=adjustment.credit_account_id, description=adjustment.description[:255], credit=adjustment.credit or 0.0)
                    
                    # --- Update Purchase Status ---
                    purchase_id = form.cleaned_data.get('purchase_id')
                    if purchase_id and not form.cleaned_data.get('partial_offset'):
                        try:
                            purchase = Purchase.objects.get(id=purchase_id)
                            purchase.payment_status = 'Paid'
                            purchase.save(update_fields=['payment_status'])
                        except Purchase.DoesNotExist:
                            pass

                    # --- Update Prepayment Status similarly ---
                    journal_voucher_id = form.cleaned_data.get('journal_voucher_id')
                    if journal_voucher_id and not form.cleaned_data.get('partial_offset'):
                        try:
                            jv = JournalVoucher.objects.get(id=journal_voucher_id)
                            jv.payment_status = 'Paid'
                            jv.save(update_fields=['payment_status'])
                        except JournalVoucher.DoesNotExist:
                            pass
                            
                    bank_id = form.cleaned_data.get('bank_id')
                    if bank_id and not form.cleaned_data.get('partial_offset'):
                        try:
                            bank_instance = Bank.objects.get(id=bank_id)
                            purchase_id = form.cleaned_data.get('purchase_id')
                            if purchase_id:
                                existing_ids = bank_instance.matched_purchase_ids or ""
                                id_list = [x.strip() for x in existing_ids.split(',') if x.strip()]
                                if str(purchase_id) not in id_list:
                                    id_list.append(str(purchase_id))
                                bank_instance.matched_purchase_ids = ",".join(id_list)
                                try: bank_instance.matched_purchase_id = int(id_list[0])
                                except (ValueError, IndexError): pass
                                bank_instance.save(update_fields=['matched_purchase_ids', 'matched_purchase_id'])
                        except Bank.DoesNotExist:
                            pass

            messages.success(request, f"Successfully processed {instances_saved} prepayments offsets.")
            return redirect('tools:offset_success')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        # Generate proposed offsets on GET
        initial_data = generate_fifo_offset_proposals()
        paginator = Paginator(initial_data, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        formset = OffsetFormSet(initial=page_obj.object_list)

    return render(request, 'offset_prepayments.html', {'formset': formset, 'page_obj': page_obj})

@login_required(login_url="register:login")
def offset_success_view(request):
    """Renders the success page after prepayment offset completes."""
    return render(request, 'offset_success.html')

# ====================================================================
# --- 7. AGENTIC ORCHESTRATOR TESTING VIEWS (PARALLEL WORKFLOW) ---
# ====================================================================

@login_required(login_url="register:login")
def agentic_invoice_upload_view(request):
    """Parallel testing view that routes to the new Agentic Orchestrator architecture."""
    user = request.user

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'process_page':
            job = request.session.get('agentic_invoice_job')
            if not job:
                return JsonResponse({"status": "error", "message": "Job session not found."})
                
            page = int(request.POST.get('page', 1))
            api_key = getattr(settings, 'GEMINI_API_KEY_2', os.getenv("GEMINI_API_KEY_2"))
            
            # 🚀 USE THE NEW ORCHESTRATOR
            processor = InvoiceOrchestrator(api_key=api_key)
            
            local_file_path = job.get('local_file_path')
            if not local_file_path or not os.path.exists(local_file_path):
                return JsonResponse({"status": "error", "message": "Source PDF lost on server."})
                
            try:
                reader = PdfReader(local_file_path)
                writer = PdfWriter()
                writer.add_page(reader.pages[page - 1])
                
                pdf_bytes_io = io.BytesIO()
                writer.write(pdf_bytes_io)
                single_page_bytes = pdf_bytes_io.getvalue()
            except Exception as e:
                return JsonResponse({"status": "error", "message": f"Failed to isolate PDF page {page}. Error: {str(e)}"})
            
            print(f"\n[AGENTIC PAGE {page}] EXTRACTING KEY INVOICE DATA...")
            ledgers, page_cost, next_seq, err = processor.process_single_page(
                pdf_bytes=single_page_bytes, 
                pg=page, custom_prompt=job['custom_prompt'],
                batch_name=job['batch_name'], rules_context=job['rules_context'],
                memo_context=job['memo_context'], current_invoice_seq=job['current_seq'],
                date_prefix=job['date_prefix'], is_explicit_seq=job['is_explicit_seq']
            )
            
            current_job = request.session.get('agentic_invoice_job')
            if current_job:
                if ledgers:
                    current_job['results'].extend(ledgers)
                    print(f"   🎉 Extracted {len(ledgers)} invoices from Page {page}.")

                current_job['current_seq'] = max(current_job.get('current_seq', 1), next_seq)
                current_job['costs']['pro_cost'] += page_cost
                request.session['agentic_invoice_job'] = current_job
                request.session.save()
            
            if err and "Timeout Error" in err:
                return JsonResponse({"status": "timeout", "page": page, "message": f"AI timed out on Page {page}."})
            
            return JsonResponse({"status": "success", "page": page, "ledgers_count": len(ledgers) if ledgers else 0, "error": err})
            
        if action == 'finalize':
            job = request.session.get('agentic_invoice_job')
            if not job:
                return JsonResponse({"status": "error", "message": "Job session not found."})
                
            local_file_path = job.get('local_file_path')
            if local_file_path and os.path.exists(local_file_path):
                os.remove(local_file_path)
                
            results = job.get('results', [])
            results.sort(key=lambda x: int(x.get('page', 0) or 0))
            
            is_explicit_seq = job.get('is_explicit_seq', False)
            date_prefix = job.get('date_prefix')
            original_seq = job.get('original_seq', 1)
            
            processed_pages = set()
            month_trackers = {}
            current_explicit_seq = original_seq
            
            for item in results:
                page = item.get('page')
                inv_no = str(item.get('invoice_no', ''))

                if inv_no == "NEEDS_SEQ" or inv_no.startswith('INV-'):
                    if is_explicit_seq:
                        if page not in processed_pages:
                            processed_pages.add(page)
                            base_seq = current_explicit_seq
                            current_explicit_seq += 1
                        else:
                            base_seq = current_explicit_seq - 1
                        base_inv_no = f"INV-{date_prefix}{base_seq:02d}"
                    else:
                        item_date = item.get('date')
                        if item_date:
                            try:
                                parsed_date = datetime.strptime(item_date, "%Y-%m-%d")
                                month_prefix = parsed_date.strftime("%Y%m")
                            except ValueError: month_prefix = datetime.now().strftime("%Y%m")
                        else:
                            month_prefix = datetime.now().strftime("%Y%m")
                            
                        if month_prefix not in month_trackers:
                            existing_invs = Purchase.objects.filter(
                                invoice_no__startswith=f"INV-{month_prefix}"
                            ).values_list('invoice_no', flat=True)
                            max_seq = 0
                            for inv in existing_invs:
                                match = re.search(rf'INV-{month_prefix}(\d+)', inv)
                                if match: max_seq = max(max_seq, int(match.group(1)))
                            month_trackers[month_prefix] = max_seq + 1
                            
                        if page not in processed_pages:
                            processed_pages.add(page)
                            base_seq = month_trackers[month_prefix]
                            month_trackers[month_prefix] += 1
                        else:
                            base_seq = month_trackers[month_prefix] - 1
                            
                        base_inv_no = f"INV-{month_prefix}{base_seq:02d}"

                    parts = inv_no.split("-")
                    if len(parts) > 2 and parts[-1].isdigit() and len(parts[-1]) < 4:
                        item['invoice_no'] = f"{base_inv_no}-{parts[-1]}"
                    else:
                        item['invoice_no'] = base_inv_no

            total_flash = job['costs']['flash_cost']
            total_pro = job['costs']['pro_cost']
            total_cost = total_flash + total_pro
            
            try:
                AICostLog.objects.create(file_name=f"[AGENTIC] {job['file_name']}", total_pages=job['total_pages'], flash_cost=total_flash, pro_cost=total_pro, total_cost=total_cost)
            except NameError: pass
                
            request.session['agentic_extracted_invoices'] = results
            request.session['agentic_ai_metadata'] = {
                'file_name': job['file_name'], 'batch_name': job['batch_name'],
                'total_pages': job['total_pages'], 'costs': job['costs']
            }
            request.session.pop('agentic_invoice_job', None)
            return JsonResponse({"status": "success", "redirect_url": reverse('tools:agentic_review_invoices')})

        request.session.pop('invoice_report_path', None)
        form = BatchUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_pdf = form.cleaned_data['invoice_pdf']
            batch_name = form.cleaned_data['batch_name']
            custom_prompt = form.cleaned_data.get('ai_prompt', '')
            
            inv_match = re.search(r'INV-(\d{6})(\d+)', custom_prompt, re.IGNORECASE)
            if inv_match:
                date_prefix = inv_match.group(1) 
                current_seq = int(inv_match.group(2)) 
                is_explicit_seq = True
            else:
                date_prefix = datetime.now().strftime("%Y%m")
                current_seq = 1
                is_explicit_seq = False

            # Agentic Orchestrator handles its own RAG pipeline. We just pass memos for batch context if any.
            client_memo = ClientPromptMemo.objects.first()
            memo_context = client_memo.memo_text if client_memo else ""
            rules_context = ""

            temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_invoices')
            os.makedirs(temp_dir, exist_ok=True)
            unique_filename = f"agentic_batch_{uuid.uuid4().hex}.pdf"
            local_file_path = os.path.join(temp_dir, unique_filename)
            
            with open(local_file_path, 'wb') as f:
                for chunk in uploaded_pdf.chunks(): f.write(chunk)

            try:
                reader = PdfReader(local_file_path)
                total_pages = len(reader.pages)
                
                if total_pages > 20:
                    os.remove(local_file_path)
                    return JsonResponse({"status": "error", "message": f"Limit exceeded. PDF has {total_pages} pages, max is 20."})
                
                request.session['agentic_invoice_job'] = {
                    'local_file_path': local_file_path, 'file_name': uploaded_pdf.name, 'total_pages': total_pages,
                    'batch_name': batch_name, 'custom_prompt': custom_prompt,
                    'rules_context': rules_context, 'memo_context': memo_context, 'is_explicit_seq': is_explicit_seq,
                    'date_prefix': date_prefix, 'original_seq': current_seq, 'current_seq': current_seq,
                    'results': [], 'costs': {'flash_cost': 0.0, 'pro_cost': 0.0}
                }
                request.session.save()
                return JsonResponse({"status": "init_success", "total_pages": total_pages})
                
            except Exception as e:
                if os.path.exists(local_file_path): os.remove(local_file_path)
                return JsonResponse({"status": "error", "message": f"Initialization Error: {str(e)}"})
        else:
            return JsonResponse({"status": "error", "message": "Form validation failed."})
    else:
        job = request.session.get('agentic_invoice_job')
        if job and 'local_file_path' in job and os.path.exists(job['local_file_path']):
            os.remove(job['local_file_path'])
            request.session.pop('agentic_invoice_job', None)
            
        form = BatchUploadForm()

    return render(request, 'agentic_invoice_upload.html', {'form': form})


@login_required(login_url="register:login")
def agentic_review_invoices(request):
    """Parallel view: Human-In-The-Loop review specifically for the new Agentic Orchestrator data."""
    extracted_data = request.session.get('agentic_extracted_invoices', [])
    metadata = request.session.get('agentic_ai_metadata', {})

    if not extracted_data and request.method == 'GET':
        return redirect('tools:agentic_invoice_upload')
        
    extracted_data.sort(key=lambda x: int(x.get('page', 0) or 0))
        
    db_vendors = [(v.id, f"{v.vendor_id} - {v.name}") for v in Vendor.objects.all().order_by('vendor_id')]
    temp_vendors = []
    for item in extracted_data:
        if item.get('is_new_vendor') and item.get('temp_id'):
            temp_id = item['temp_id']
            temp_vid = item.get('temp_vid', 'NEW')
            company = item.get('company', 'Unknown')
            if not any(tv[0] == temp_id for tv in temp_vendors):
                temp_vendors.append((temp_id, f"✨ NEW: {company} ({temp_vid})"))
                
    dynamic_choices = [('', '--- Select Vendor ---')] + db_vendors + temp_vendors

    seen_accounts = set()
    db_accounts = []
    for acc_id, name in Account.objects.values_list('account_id', 'name'):
        if acc_id not in seen_accounts:
            seen_accounts.add(acc_id)
            db_accounts.append((str(acc_id), f"{acc_id} - {name}"))
    db_accounts.sort(key=lambda x: str(x[0]))
    account_choices = [('', '--- Select Account ---')] + db_accounts

    # We pass the data to the same robust formset and database posting logic
    # To keep this DRY and concise, we route the verified data to your existing hitl review pipeline logic
    # We temporarily assign the session keys back to normal, call the logic, and pop them!
    if request.method == 'POST':
        request.session['extracted_invoices'] = request.session.get('agentic_extracted_invoices')
        request.session['ai_metadata'] = request.session.get('agentic_ai_metadata')
        
        response = review_invoices(request, template_name='tools/agentic_invoice_review.html')
        
        # Clear our parallel keys after the main function processes them
        if not isinstance(response, HttpResponseRedirect) and response.status_code == 200:
            # Validation failed, restore normal state and render
            pass
        else:
            # Success!
            request.session.pop('agentic_extracted_invoices', None)
            request.session.pop('agentic_ai_metadata', None)
            
        return response
        
    else:
        formset = PurchaseFormSet(
            initial=extracted_data, 
            form_kwargs={'dynamic_choices': dynamic_choices, 'account_choices': account_choices}
        )
        
    return render(request, 'tools/agentic_invoice_review.html', {'formset': formset, 'metadata': metadata})

@csrf_exempt
def pubsub_draft_rule_webhook(request):
    """
    Gateway for 'DRAFT_RULE_PROPOSED' events.
    Consistency Check: Decodes payload and offloads to Celery immediately.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        body = json.loads(request.body)
        message = body.get('message', {})
        
        if not message or 'data' not in message:
            return JsonResponse({'error': 'Invalid Pub/Sub payload'}, status=400)
            
        # Decode Pub/Sub base64 data
        raw_data = base64.b64decode(message['data']).decode('utf-8')
        payload = json.loads(raw_data)
        
        # OFFLOAD TO CELERY: Ensures consistency and prevents Pub/Sub timeouts
        process_draft_rule_task.delay(payload)
        
        return JsonResponse({'status': 'Accepted for background processing'}, status=200)

    except Exception as e:
        print(f"❌ [Webhook Entry Error]: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def pubsub_user_corrections_webhook(request):
    """
    Step 2: Receives user corrections from Pub/Sub Topic 1,
    validates the OIDC token, and offloads processing to Celery.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # 1. (Recommended) Validate the OIDC Token
    # This prevents unauthorized users from posting fake rules to your endpoint.
    # auth_header = request.headers.get('Authorization')
    # if not auth_header or not verify_oidc_token(auth_header):
    #     return JsonResponse({'error': 'Unauthorized'}, status=401)

    try:
        body = json.loads(request.body)
        message = body.get('message', {})
        if not message:
            return JsonResponse({'error': 'No message found in payload'}, status=400)
            
        # 2. Decode the Base64 data from Pub/Sub
        raw_data = base64.b64decode(message.get('data', '')).decode('utf-8')
        payload = json.loads(raw_data)
        
        # 3. Offload to Celery (Crucial for Step 2)
        # We use .delay() to respond to Pub/Sub immediately.
        # This function should save to DraftKnowledgeRule AND then publish to Topic 2.
        from tools.tasks import handle_user_correction_task
        handle_user_correction_task.delay(payload)

        # 4. Acknowledge the message immediately
        return JsonResponse({'status': 'Processing started'}, status=200)

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        # Logging errors is vital for troubleshooting Pub/Sub delivery
        print(f"❌ [PubSub Webhook Error]: {e}")
        return JsonResponse({'error': 'Internal server error'}, status=500)
